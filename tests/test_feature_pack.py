from __future__ import annotations

import base64
import csv
import io
import json
from pathlib import Path
import re
import sys
import time

from openpyxl import load_workbook

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app import create_app
from service_manager.db import get_db, inserted_id, transaction
from service_manager.audit import append_audit_event
from service_manager.crypto import account_field_aad, account_password_aad, encrypt_secret


KEY = base64.b64encode(b"f" * 32).decode("ascii")
ADMIN_PASSWORD = "admin-password-0123456789"


@pytest.fixture()
def app(tmp_path: Path):
    return create_app(
        {
            "TESTING": True,
            "PROPAGATE_EXCEPTIONS": False,
            "DATABASE_PATH": str(tmp_path / "feature-pack.db"),
            "DATA_KEY_V1": KEY,
            "AUDIT_KEY_V1": KEY,
            "SECRET_KEY": "feature-pack-session-secret",
            "WTF_CSRF_ENABLED": False,
            "ADMIN_USERNAME": "admin",
            "ADMIN_PASSWORD": ADMIN_PASSWORD,
        }
    )


@pytest.fixture()
def client(app):
    return app.test_client()


def login_admin(client) -> None:
    response = client.post("/login", data={"username": "admin", "password": ADMIN_PASSWORD})
    assert response.status_code == 302

def login_operator(app, client) -> None:
    with app.app_context():
        conn = get_db()
        admin = conn.execute("SELECT id FROM users WHERE username='admin'").fetchone()
        stamp = conn.execute("SELECT created_at FROM users WHERE id=?", (admin["id"],)).fetchone()[0]
        from service_manager.crypto import hash_password

        conn.execute(
            "INSERT INTO users (username, password_hash, role, is_active, must_change_password, created_at, updated_at) VALUES (?, ?, 'operador', 1, 0, ?, ?)",
            ("operator", hash_password("operator-password-012345"), stamp, stamp),
        )
        conn.commit()
    response = client.post("/login", data={"username": "operator", "password": "operator-password-012345"})
    assert response.status_code == 302


def enable_rotation(app) -> None:
    with app.app_context():
        conn = get_db()
        conn.execute("INSERT OR REPLACE INTO app_settings (key, value) VALUES ('rotation_enabled', '1')")
        conn.commit()


from datetime import date
from service_manager.routes import _parse_rotation_days, _parse_rotation_due_at, _rotation_state


def test_parse_rotation_days_semantics():
    # Absent/blank -> inherit/clear.
    assert _parse_rotation_days(None) == (True, None)
    assert _parse_rotation_days("") == (True, None)
    assert _parse_rotation_days("   ") == (True, None)
    # Valid ASCII decimals within range.
    assert _parse_rotation_days("1") == (True, 1)
    assert _parse_rotation_days("90") == (True, 90)
    assert _parse_rotation_days("3650") == (True, 3650)
    assert _parse_rotation_days(" 30 ") == (False, None)  # Whitespace-wrapped non-blank is invalid.
    # Out of range and non-ASCII-decimal inputs are invalid.
    assert _parse_rotation_days("0") == (False, None)
    assert _parse_rotation_days("3651") == (False, None)
    assert _parse_rotation_days("-5") == (False, None)
    assert _parse_rotation_days("12.5") == (False, None)
    assert _parse_rotation_days("abc") == (False, None)
    assert _parse_rotation_days("٩") == (False, None)  # Arabic-Indic digit rejected.
    assert _parse_rotation_days("1_0") == (False, None)


def test_parse_rotation_due_at_requires_exact_canonical_iso():
    assert _parse_rotation_due_at(None) == (True, None)
    assert _parse_rotation_due_at("") == (True, None)
    assert _parse_rotation_due_at("2026-07-22") == (True, "2026-07-22")
    assert _parse_rotation_due_at(" 2026-07-22 ") == (False, None)  # Whitespace-wrapped non-blank is invalid.
    # Non-canonical or malformed inputs rejected.
    assert _parse_rotation_due_at("2026-7-2") == (False, None)
    assert _parse_rotation_due_at("2026/07/22") == (False, None)
    assert _parse_rotation_due_at("22-07-2026") == (False, None)
    assert _parse_rotation_due_at("2026-13-01") == (False, None)
    assert _parse_rotation_due_at("not-a-date") == (False, None)


def test_rotation_state_unknown_and_no_policy():
    today = date(2026, 7, 22)
    # Absent password timestamp with no policy -> unknown (existing accounts).
    assert _rotation_state(None, None, None, None, today=today)["state"] == "unknown"
    # Absent password timestamp even with a policy -> unknown.
    assert _rotation_state(None, 30, None, None, today=today)["state"] == "unknown"
    # Malformed / naive timestamps fail closed to unknown.
    assert _rotation_state("not-a-timestamp", None, None, 30, today=today)["state"] == "unknown"
    assert _rotation_state("2026-07-01T00:00:00", None, None, 30, today=today)["state"] == "unknown"
    # Valid tz-aware timestamp but no effective policy -> no_policy.
    result = _rotation_state("2026-07-01T00:00:00+00:00", None, None, None, today=today)
    assert result["state"] == "no_policy"
    assert result["due_at"] is None and result["days_remaining"] is None


def test_rotation_state_precedence_and_boundaries():
    today = date(2026, 7, 22)
    # Explicit due-date override wins even when password history is unknown.
    overdue = _rotation_state(None, None, "2026-07-20", 30, today=today)
    assert overdue["state"] == "overdue" and overdue["days_remaining"] == -2
    # Account-level days override the service default.
    result = _rotation_state("2026-07-01T00:00:00+00:00", 10, None, 365, today=today)
    assert result["effective_days"] == 10 and result["due_at"] == "2026-07-11"
    assert result["state"] == "overdue"
    # Service default applies when account days is NULL.
    svc = _rotation_state("2026-07-01T00:00:00+00:00", None, None, 30, today=today)
    assert svc["effective_days"] == 30 and svc["due_at"] == "2026-07-31"
    # Seven-day boundary: exactly 7 days remaining is due_soon; 8 is current.
    due_soon = _rotation_state("2026-07-01T00:00:00+00:00", None, None, 28, today=today)  # due 2026-07-29 -> 7 days
    assert due_soon["days_remaining"] == 7 and due_soon["state"] == "due_soon"
    current = _rotation_state("2026-07-01T00:00:00+00:00", None, None, 29, today=today)  # due 2026-07-30 -> 8 days
    assert current["days_remaining"] == 8 and current["state"] == "current"
    # Zero days remaining (due today) is due_soon, not overdue.
    zero = _rotation_state(None, None, "2026-07-22", None, today=today)
    assert zero["days_remaining"] == 0 and zero["state"] == "due_soon"
    # days_remaining == -1 is the immediate overdue threshold.
    minus_one = _rotation_state(None, None, "2026-07-21", None, today=today)
    assert minus_one["days_remaining"] == -1 and minus_one["state"] == "overdue"
    # A malformed explicit due override fails closed to unknown.
    assert _rotation_state("2026-07-01T00:00:00+00:00", None, "2026-13-40", 30, today=today)["state"] == "unknown"
    # Non-UTC tz-aware timestamps normalize to UTC date before adding the interval.
    tz = _rotation_state("2026-06-30T23:00:00-05:00", None, None, 30, today=today)  # = 2026-07-01T04:00Z
    assert tz["due_at"] == "2026-07-31"

def test_admin_users_renders_management_page(client):
    login_admin(client)

    response = client.get("/admin/users")

    assert response.status_code == 200
    assert response.headers["Cache-Control"] == "no-store, private"
    body = response.get_data(as_text=True)
    assert "data-admin-create" in body
    assert "data-admin-role" in body
    assert "data-admin-active" in body
    assert "admin" in body

    client.post("/logout")
    login_operator(client.application, client)
    assert client.get("/admin/users").status_code == 403


def test_admin_users_redirects_anonymous_user(client):
    response = client.get("/admin/users")

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/login")


def test_index_renders_url_state_hooks(client):
    login_admin(client)
    with client.application.app_context():
        conn = get_db()
        service_id = conn.execute("INSERT INTO services (name) VALUES ('Email')").lastrowid
        account_id = conn.execute(
            "INSERT INTO accounts (email, password_ciphertext, password_nonce, password_key_version) VALUES ('hook@example.test', ?, ?, 1)",
            (b"x", b"0" * 12),
        ).lastrowid
        conn.execute("INSERT INTO account_service (account_id, service_id, status, registered) VALUES (?, ?, 'ativo', 1)", (account_id, service_id))
        conn.commit()

    response = client.get("/")

    assert response.status_code == 200
    body = response.get_data(as_text=True)
    assert 'id="account-filter"' in body
    assert 'id="filter-status"' in body
    assert 'id="filter-registered"' in body
    assert 'id="share-view"' in body

    assert "data-row-select" in body


def test_index_renders_bulk_field_and_typed_delete_hooks(app, client):
    login_admin(client)
    service_id, account_ids, _ = seed_bulk_data(app)
    empty_service_id = None
    with app.app_context():
        conn = get_db()
        conn.execute("INSERT INTO custom_fields (service_id, name) VALUES (?, 'PIN')", (service_id,))
        empty_service_id = inserted_id(conn.execute("INSERT INTO services (name) VALUES ('SemCampos')"))
        conn.commit()

    body = client.get(f"/?service={service_id}").get_data(as_text=True)

    assert 'id="bulk-field-id"' in body
    assert 'id="bulk-field-value"' in body
    assert 'id="bulk-apply-field"' in body
    assert 'id="delete-confirm-dialog"' in body
    assert 'id="delete-confirm-input"' in body
    assert 'id="bulk-add-field"' in body
    assert 'id="bulk-field-dialog"' in body
    assert 'data-bulk-field-form' in body
    assert 'id="bulk-field-name"' in body
    assert 'O campo será criado vazio para preenchimento individual.' in body
    assert '/accounts/bulk/field/add' in body

    # Button is available even on a service without any custom field yet.
    empty_body = client.get(f"/?service={empty_service_id}").get_data(as_text=True)
    assert 'id="bulk-add-field"' in empty_body
    assert 'id="bulk-field-dialog"' in empty_body
    assert 'id="bulk-field-id"' not in empty_body

    # Viewers get neither the button nor the dialog; editors do.
    with app.app_context():
        conn = get_db()
        from service_manager.crypto import hash_password
        stamp = conn.execute("SELECT created_at FROM users WHERE username='admin'").fetchone()[0]
        for username, role in (("viewer-u", "viewer"), ("editor-u", "editor")):
            uid = inserted_id(conn.execute(
                "INSERT INTO users (username, password_hash, role, is_active, must_change_password, created_at, updated_at) VALUES (?, ?, 'operador', 1, 0, ?, ?)",
                (username, hash_password("member-password-012345"), stamp, stamp),
            ))
            conn.execute(
                "INSERT INTO service_members (user_id, service_id, role, created_at) VALUES (?, ?, ?, '2026-01-01T00:00:00+00:00')",
                (uid, service_id, role),
            )
        conn.commit()

    client.post("/logout")
    assert client.post("/login", data={"username": "viewer-u", "password": "member-password-012345"}).status_code == 302
    viewer_body = client.get(f"/?service={service_id}").get_data(as_text=True)
    assert 'id="bulk-add-field"' not in viewer_body
    assert 'id="bulk-field-dialog"' not in viewer_body

    client.post("/logout")
    assert client.post("/login", data={"username": "editor-u", "password": "member-password-012345"}).status_code == 302
    editor_body = client.get(f"/?service={service_id}").get_data(as_text=True)
    assert 'id="bulk-add-field"' in editor_body
    assert 'id="bulk-field-dialog"' in editor_body

def _make_member(app, username, role, service_id, password="member-password-012345") -> int:
    with app.app_context():
        conn = get_db()
        from service_manager.crypto import hash_password

        stamp = conn.execute("SELECT created_at FROM users WHERE username='admin'").fetchone()[0]
        uid = inserted_id(conn.execute(
            "INSERT INTO users (username, password_hash, role, is_active, must_change_password, created_at, updated_at) VALUES (?, ?, 'operador', 1, 0, ?, ?)",
            (username, hash_password(password), stamp, stamp),
        ))
        conn.execute(
            "INSERT INTO service_members (user_id, service_id, role, created_at) VALUES (?, ?, ?, '2026-01-01T00:00:00+00:00')",
            (uid, service_id, role),
        )
        conn.commit()
    return uid


def reauth(client, password=ADMIN_PASSWORD) -> None:
    assert client.post("/reauth", data={"password": password}).status_code == 204


def seed_export_data(app) -> tuple[int, list[str], list[list[str]]]:
    """Seed a service with two accounts and three custom fields.

    Returns ``(service_id, header, rows)`` describing the exact export matrix.
    """
    with app.app_context():
        conn = get_db()
        service_id = inserted_id(conn.execute("INSERT INTO services (name) VALUES ('Export')"))
        # Account 1: password and account email both begin with a formula char.
        acc1 = inserted_id(conn.execute(
            "INSERT INTO accounts (email, password_ciphertext, password_nonce, password_key_version) VALUES (?, ?, ?, 1)",
            ("=formula@example.test", b"", b"0" * 12),
        ))
        pw1 = encrypt_secret("=secret-pw-1", aad=account_password_aad(acc1))
        conn.execute("UPDATE accounts SET password_ciphertext=?, password_nonce=? WHERE id=?", (pw1.ciphertext, pw1.nonce, acc1))
        conn.execute("INSERT INTO account_service (account_id, service_id, status, registered) VALUES (?, ?, 'ativo', 1)", (acc1, service_id))
        # Account 2: no field associations.
        acc2 = inserted_id(conn.execute(
            "INSERT INTO accounts (email, password_ciphertext, password_nonce, password_key_version) VALUES (?, ?, ?, 1)",
            ("second@example.test", b"", b"0" * 12),
        ))
        pw2 = encrypt_secret("plain-pw-2", aad=account_password_aad(acc2))
        conn.execute("UPDATE accounts SET password_ciphertext=?, password_nonce=? WHERE id=?", (pw2.ciphertext, pw2.nonce, acc2))
        conn.execute("INSERT INTO account_service (account_id, service_id, status, registered) VALUES (?, ?, 'nunca', 0)", (acc2, service_id))
        # Three fields; ORDER BY name (BINARY) => 'API key', 'Vazio', 'email'.
        for name, value in (("API key", "key-value-123"), ("email", "=EMAIL_FORMULA"), ("Vazio", "")):
            fid = inserted_id(conn.execute("INSERT INTO custom_fields (service_id, name) VALUES (?, ?)", (service_id, name)))
            enc = encrypt_secret(value, aad=account_field_aad(acc1, fid))
            conn.execute(
                "INSERT INTO field_values (field_id, account_id, value_ciphertext, value_nonce, value_key_version) VALUES (?, ?, ?, ?, 1)",
                (fid, acc1, enc.ciphertext, enc.nonce),
            )
        # A filled field on a different service must never leak into this export.
        other_id = inserted_id(conn.execute("INSERT INTO services (name) VALUES ('Other')"))
        other_field = inserted_id(conn.execute("INSERT INTO custom_fields (service_id, name) VALUES (?, 'Segredo alheio')", (other_id,)))
        conn.execute("INSERT INTO account_service (account_id, service_id, status, registered) VALUES (?, ?, 'ativo', 0)", (acc1, other_id))
        other_val = encrypt_secret("outro-servico-secreto", aad=account_field_aad(acc1, other_field))
        conn.execute(
            "INSERT INTO field_values (field_id, account_id, value_ciphertext, value_nonce, value_key_version) VALUES (?, ?, ?, ?, 1)",
            (other_field, acc1, other_val.ciphertext, other_val.nonce),
        )
        conn.commit()
    header = ["email", "password", "status", "cadastrada", "campo:API key", "campo:Vazio", "campo:email"]
    rows = [
        ["'=formula@example.test", "'=secret-pw-1", "ativo", "sim", "key-value-123", "", "'=EMAIL_FORMULA"],
        ["second@example.test", "plain-pw-2", "nunca", "não", "", "", ""],
    ]
    return service_id, header, rows


def test_csv_export_reveals_matrix_only_after_reauth(app, client):
    login_admin(client)
    service_id, header, rows = seed_export_data(app)
    reauth(client)

    response = client.get(f"/export.csv?service={service_id}")

    assert response.status_code == 200
    assert response.mimetype == "text/csv"
    disp = response.headers["Content-Disposition"]
    assert f"filename=contas_export_{service_id}_" in disp and disp.endswith(".csv")
    assert response.headers["Cache-Control"] == "no-store, private"
    assert response.headers["Pragma"] == "no-cache"
    text = response.get_data(as_text=True)
    assert text.startswith("\ufeff")
    reader = list(csv.reader(io.StringIO(text.lstrip("\ufeff"))))
    assert reader[0] == header
    assert reader[1:] == rows
    # Nothing from the other service leaks.
    assert "Segredo alheio" not in text
    assert "outro-servico-secreto" not in text


def test_xlsx_export_matches_csv_matrix_after_reauth(app, client):
    login_admin(client)
    service_id, header, rows = seed_export_data(app)
    reauth(client)

    response = client.get(f"/export.xlsx?service={service_id}")

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    disp = response.headers["Content-Disposition"]
    assert f"filename=contas_export_{service_id}_" in disp and disp.endswith(".xlsx")
    assert response.headers["Cache-Control"] == "no-store, private"
    assert response.headers["Pragma"] == "no-cache"
    workbook = load_workbook(io.BytesIO(response.get_data()), read_only=True, data_only=True)
    sheet = workbook.active
    assert sheet is not None
    matrix = []
    for row in sheet.iter_rows(values_only=True):
        cells = ["" if cell is None else str(cell) for cell in row]
        cells += [""] * (len(header) - len(cells))
        matrix.append(cells)
    workbook.close()
    assert matrix[0] == header
    assert matrix[1:] == rows


@pytest.mark.parametrize("ext", ["csv", "xlsx"])
def test_export_forbidden_below_service_admin(app, client, ext):
    service_id, _, _ = seed_export_data(app)
    # Plain operator with no membership.
    login_operator(app, client)
    with client.session_transaction() as session:
        session["reauthenticated_at"] = time.time()
    assert client.get(f"/export.{ext}?service={service_id}").status_code == 403
    client.post("/logout")
    # Viewer and editor members remain below service_admin.
    _make_member(app, "viewer-x", "viewer", service_id)
    _make_member(app, "editor-x", "editor", service_id)
    for username in ("viewer-x", "editor-x"):
        assert client.post("/login", data={"username": username, "password": "member-password-012345"}).status_code == 302
        with client.session_transaction() as session:
            session["reauthenticated_at"] = time.time()
        assert client.get(f"/export.{ext}?service={service_id}").status_code == 403
        client.post("/logout")
    with app.app_context():
        assert get_db().execute("SELECT COUNT(*) FROM audit_events WHERE action='accounts.exported'").fetchone()[0] == 0


@pytest.mark.parametrize("ext", ["csv", "xlsx"])
def test_admin_export_requires_recent_reauth(app, client, ext):
    login_admin(client)
    service_id, _, _ = seed_export_data(app)
    # Admin without recent reauth is rejected and nothing is audited.
    assert client.get(f"/export.{ext}?service={service_id}").status_code == 403
    with app.app_context():
        assert get_db().execute("SELECT COUNT(*) FROM audit_events WHERE action='accounts.exported'").fetchone()[0] == 0
    reauth(client)
    response = client.get(f"/export.{ext}?service={service_id}")
    assert response.status_code == 200
    with app.app_context():
        events = get_db().execute("SELECT metadata_json FROM audit_events WHERE action='accounts.exported'").fetchall()
    assert len(events) == 1
    assert json.loads(events[0]["metadata_json"]) == {"rows": 2, "format": ext}


@pytest.mark.parametrize("ext", ["csv", "xlsx"])
def test_service_admin_member_exports_after_reauth(app, client, ext):
    service_id, _, _ = seed_export_data(app)
    _make_member(app, "svc-admin", "service_admin", service_id)
    assert client.post("/login", data={"username": "svc-admin", "password": "member-password-012345"}).status_code == 302
    # No reauth yet: rejected and unaudited.
    assert client.get(f"/export.{ext}?service={service_id}").status_code == 403
    with app.app_context():
        assert get_db().execute("SELECT COUNT(*) FROM audit_events WHERE action='accounts.exported'").fetchone()[0] == 0
    reauth(client, password="member-password-012345")
    response = client.get(f"/export.{ext}?service={service_id}")
    assert response.status_code == 200
    with app.app_context():
        events = get_db().execute("SELECT metadata_json FROM audit_events WHERE action='accounts.exported'").fetchall()
    assert len(events) == 1
    assert json.loads(events[0]["metadata_json"]) == {"rows": 2, "format": ext}


def test_export_filename_slugs_name_with_id_and_utc_timestamp(app, client):
    login_admin(client)
    with app.app_context():
        conn = get_db()
        service_id = inserted_id(conn.execute("INSERT INTO services (name) VALUES ('Serviço Especial!!')"))
        conn.commit()
    reauth(client)

    csv_resp = client.get(f"/export.csv?service={service_id}")
    xlsx_resp = client.get(f"/export.xlsx?service={service_id}")

    for resp, ext in ((csv_resp, "csv"), (xlsx_resp, "xlsx")):
        assert resp.status_code == 200
        disp = resp.headers["Content-Disposition"]
        assert re.search(rf"filename=contas_servico_especial_{service_id}_\d{{8}}T\d{{6}}Z\.{ext}", disp), disp


def test_export_limit_rejects_over_limit_without_audit(app, client, monkeypatch):
    import service_manager.routes as routes_module

    assert routes_module._EXPORT_LIMIT == 10000
    monkeypatch.setattr(routes_module, "_EXPORT_LIMIT", 2)
    login_admin(client)
    reauth(client)
    account_ids: list[int] = []
    with app.app_context():
        conn = get_db()
        service_id = inserted_id(conn.execute("INSERT INTO services (name) VALUES ('Big')"))
        for n in range(3):
            aid = inserted_id(conn.execute(
                "INSERT INTO accounts (email, password_ciphertext, password_nonce, password_key_version) VALUES (?, ?, ?, 1)",
                (f"user{n}@x.test", b"", b"0" * 12),
            ))
            envelope = encrypt_secret(f"pw-{n}", aad=account_password_aad(aid))
            conn.execute("UPDATE accounts SET password_ciphertext=?, password_nonce=? WHERE id=?", (envelope.ciphertext, envelope.nonce, aid))
            conn.execute("INSERT INTO account_service (account_id, service_id, status, registered) VALUES (?, ?, 'nunca', 0)", (aid, service_id))
            account_ids.append(aid)
        conn.commit()

    over = client.get(f"/export.csv?service={service_id}")
    assert over.status_code == 413
    with app.app_context():
        assert get_db().execute("SELECT COUNT(*) FROM audit_events WHERE action='accounts.exported'").fetchone()[0] == 0

    with app.app_context():
        conn = get_db()
        conn.execute("DELETE FROM account_service WHERE account_id=?", (account_ids[-1],))
        conn.commit()
    exact = client.get(f"/export.csv?service={service_id}")
    assert exact.status_code == 200
    reader = list(csv.reader(io.StringIO(exact.get_data(as_text=True).lstrip("\ufeff"))))
    assert len(reader) == 3  # header + two accounts
    with app.app_context():
        assert get_db().execute("SELECT COUNT(*) FROM audit_events WHERE action='accounts.exported'").fetchone()[0] == 1


@pytest.mark.parametrize("target", ["password_ciphertext", "password_nonce", "field_ciphertext", "field_nonce"])
def test_export_corrupt_envelope_fails_without_audit(app, client, target):
    login_admin(client)
    service_id, _, _ = seed_export_data(app)
    reauth(client)
    with app.app_context():
        conn = get_db()
        acc = conn.execute("SELECT id FROM accounts WHERE email='=formula@example.test'").fetchone()["id"]
        if target == "password_ciphertext":
            conn.execute("UPDATE accounts SET password_ciphertext=? WHERE id=?", (b"\x00tampered-ciphertext", acc))
        elif target == "password_nonce":
            conn.execute("UPDATE accounts SET password_nonce=? WHERE id=?", (b"z" * 12, acc))
        elif target == "field_ciphertext":
            conn.execute("UPDATE field_values SET value_ciphertext=? WHERE account_id=?", (b"\x00tampered-ciphertext", acc))
        else:
            conn.execute("UPDATE field_values SET value_nonce=? WHERE account_id=?", (b"z" * 12, acc))
        conn.commit()

    response = client.get(f"/export.csv?service={service_id}")
    assert response.status_code == 500
    with app.app_context():
        assert get_db().execute("SELECT COUNT(*) FROM audit_events WHERE action='accounts.exported'").fetchone()[0] == 0


def seed_bulk_data(app) -> tuple[int, list[int], int]:
    with app.app_context():
        conn = get_db()
        service_id = inserted_id(conn.execute("INSERT INTO services (name) VALUES ('Bulk')"))
        other_service_id = inserted_id(conn.execute("INSERT INTO services (name) VALUES ('Other')"))
        account_ids: list[int] = []
        for email in ("bulk-one@example.test", "bulk-two@example.test"):
            account_id = inserted_id(conn.execute(
                "INSERT INTO accounts (email, password_ciphertext, password_nonce, password_key_version) VALUES (?, ?, ?, 1)",
                (email, b"x", b"0" * 12),
            ))
            conn.execute("INSERT INTO account_service (account_id, service_id, status, registered) VALUES (?, ?, 'nunca', 0)", (account_id, service_id))
            account_ids.append(account_id)
        foreign_id = inserted_id(conn.execute(
            "INSERT INTO accounts (email, password_ciphertext, password_nonce, password_key_version) VALUES ('foreign@example.test', ?, ?, 1)",
            (b"x", b"0" * 12),
        ))
        conn.execute("INSERT INTO account_service (account_id, service_id, status, registered) VALUES (?, ?, 'nunca', 0)", (foreign_id, other_service_id))
        conn.commit()
        return service_id, account_ids, foreign_id


def test_bulk_status_updates_accounts_and_audits(app, client):
    login_admin(client)
    service_id, account_ids, _ = seed_bulk_data(app)

    response = client.post("/accounts/bulk/status", data={"service_id": service_id, "account_ids": account_ids, "status": "ativo"})

    assert response.status_code == 302
    assert "ok=bulk_updated" in response.headers["Location"]
    with app.app_context():
        conn = get_db()
        assert [row["status"] for row in conn.execute("SELECT status FROM account_service WHERE service_id=? ORDER BY account_id", (service_id,))] == ["ativo", "ativo"]
        event = conn.execute("SELECT metadata_json FROM audit_events WHERE action='accounts.bulk_status' ORDER BY id DESC LIMIT 1").fetchone()
        assert '"count":2' in event["metadata_json"]


def test_bulk_rejects_foreign_account_and_over_limit(app, client):
    login_admin(client)
    service_id, _, foreign_id = seed_bulk_data(app)

    assert client.post("/accounts/bulk/status", data={"service_id": service_id, "account_ids": [foreign_id], "status": "ativo"}).status_code == 404
    assert client.post("/accounts/bulk/status", data={"service_id": service_id, "account_ids": list(range(1, 202)), "status": "ativo"}).status_code == 400


def test_bulk_delete_requires_admin_and_removes_accounts(app, client):
    service_id, account_ids, _ = seed_bulk_data(app)
    login_operator(app, client)
    assert client.post("/accounts/bulk/delete", data={"service_id": service_id, "account_ids": account_ids, "confirmation_count": str(len(account_ids))}).status_code == 403
    client.post("/logout")
    login_admin(client)

    response = client.post("/accounts/bulk/delete", data={"service_id": service_id, "account_ids": account_ids, "confirmation_count": str(len(account_ids))})

    assert response.status_code == 302
    assert "ok=bulk_deleted" in response.headers["Location"]
    with app.app_context():
        conn = get_db()
        assert conn.execute("SELECT COUNT(*) FROM accounts WHERE id IN (?, ?)", account_ids).fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM audit_events WHERE action='accounts.bulk_deleted'").fetchone()[0] == 1



def test_bulk_delete_requires_matching_confirmation_count(app, client):
    service_id, account_ids, _ = seed_bulk_data(app)
    login_admin(client)

    absent = client.post("/accounts/bulk/delete", data={"service_id": service_id, "account_ids": account_ids})
    mismatch = client.post("/accounts/bulk/delete", data={"service_id": service_id, "account_ids": account_ids, "confirmation_count": "1"})

    assert absent.status_code == 400
    assert mismatch.status_code == 400
    with app.app_context():
        conn = get_db()
        assert conn.execute("SELECT COUNT(*) FROM accounts WHERE id IN (?, ?)", account_ids).fetchone()[0] == 2
        assert conn.execute("SELECT COUNT(*) FROM audit_events WHERE action='accounts.bulk_deleted'").fetchone()[0] == 0


def test_bulk_field_encrypts_value_on_selected_accounts(app, client):
    login_admin(client)
    service_id, account_ids, foreign_id = seed_bulk_data(app)
    with app.app_context():
        conn = get_db()
        field_id = inserted_id(conn.execute("INSERT INTO custom_fields (service_id, name) VALUES (?, 'PIN')", (service_id,)))
        conn.commit()

    response = client.post(
        "/accounts/bulk/field",
        data={"service_id": service_id, "account_ids": account_ids, "field_id": field_id, "field_value": "1234-secret"},
    )

    assert response.status_code == 302
    with app.app_context():
        conn = get_db()
        from service_manager.crypto import decrypt_secret, EncryptedValue
        rows = conn.execute("SELECT account_id, value_ciphertext, value_nonce, value_key_version FROM field_values WHERE field_id=? ORDER BY account_id", (field_id,)).fetchall()
        assert [r["account_id"] for r in rows] == sorted(account_ids)
        for r in rows:
            assert decrypt_secret(EncryptedValue(r["value_ciphertext"], r["value_nonce"], r["value_key_version"]), aad=account_field_aad(r["account_id"], field_id)) == "1234-secret"
        event = conn.execute("SELECT metadata_json FROM audit_events WHERE action='accounts.bulk_field' ORDER BY id DESC LIMIT 1").fetchone()
        assert "1234-secret" not in event["metadata_json"]
        assert '"count":2' in event["metadata_json"] or '"count": 2' in event["metadata_json"]


def test_bulk_field_rejects_foreign_field_blank_and_oversized(app, client):
    login_admin(client)
    service_id, account_ids, _ = seed_bulk_data(app)
    with app.app_context():
        conn = get_db()
        other_field = inserted_id(conn.execute("INSERT INTO custom_fields (service_id, name) VALUES ((SELECT id FROM services WHERE name='Other'), 'X')"))
        service_field = inserted_id(conn.execute("INSERT INTO custom_fields (service_id, name) VALUES (?, 'Y')", (service_id,)))
        conn.commit()

    foreign = client.post("/accounts/bulk/field", data={"service_id": service_id, "account_ids": account_ids, "field_id": other_field, "field_value": "v"})
    blank = client.post("/accounts/bulk/field", data={"service_id": service_id, "account_ids": account_ids, "field_id": service_field, "field_value": ""})
    oversized = client.post("/accounts/bulk/field", data={"service_id": service_id, "account_ids": account_ids, "field_id": service_field, "field_value": "x" * 4097})

    assert foreign.status_code == 404
    assert blank.status_code == 400
    assert oversized.status_code == 400

def test_bulk_field_add_creates_empty_fields_for_individual_fill(app, client):
    login_admin(client)
    service_id, account_ids, _ = seed_bulk_data(app)

    response = client.post(
        "/accounts/bulk/field/add",
        data={"service_id": service_id, "account_ids": account_ids, "field_name": "Identificador externo"},
    )

    assert response.status_code == 302
    assert "ok=bulk_field_created" in response.headers["Location"]
    with app.app_context():
        conn = get_db()
        from service_manager.crypto import decrypt_secret, EncryptedValue
        fields = conn.execute("SELECT id, name FROM custom_fields WHERE service_id=?", (service_id,)).fetchall()
        assert [f["name"] for f in fields] == ["Identificador externo"]
        field_id = fields[0]["id"]
        rows = conn.execute("SELECT account_id, value_ciphertext, value_nonce, value_key_version FROM field_values WHERE field_id=? ORDER BY account_id", (field_id,)).fetchall()
        assert [r["account_id"] for r in rows] == sorted(account_ids)
        for r in rows:
            assert decrypt_secret(EncryptedValue(r["value_ciphertext"], r["value_nonce"], r["value_key_version"]), aad=account_field_aad(r["account_id"], field_id)) == ""
        event = conn.execute("SELECT metadata_json FROM audit_events WHERE action='accounts.bulk_field_created' ORDER BY id DESC LIMIT 1").fetchone()
        assert "Identificador externo" not in event["metadata_json"]
        assert ('"count":2' in event["metadata_json"]) or ('"count": 2' in event["metadata_json"])
        assert ('"created_count":2' in event["metadata_json"]) or ('"created_count": 2' in event["metadata_json"])

    body = client.get(f"/?service={service_id}").get_data(as_text=True)
    assert body.count("Identificador externo") >= 2
    assert body.count('name="value" value=""') >= 2

    lower, higher = sorted(account_ids)
    assert client.post(f"/field/update/{field_id}/{lower}", data={"service_id": service_id, "value": "ABC"}).status_code == 302
    assert client.post(f"/field/update/{field_id}/{higher}", data={"service_id": service_id, "value": "XYZ"}).status_code == 302
    filled = client.get(f"/?service={service_id}").get_data(as_text=True)
    assert 'value="ABC"' in filled
    assert 'value="XYZ"' in filled


def test_bulk_field_add_reuses_field_and_preserves_filled_values(app, client):
    login_admin(client)
    service_id, account_ids, _ = seed_bulk_data(app)
    filled_id, empty_id = sorted(account_ids)
    with app.app_context():
        conn = get_db()
        field_id = inserted_id(conn.execute("INSERT INTO custom_fields (service_id, name) VALUES (?, 'Existente')", (service_id,)))
        encrypted = encrypt_secret("mantido", aad=account_field_aad(filled_id, field_id))
        conn.execute(
            "INSERT INTO field_values (field_id, account_id, value_ciphertext, value_nonce, value_key_version) VALUES (?, ?, ?, ?, ?)",
            (field_id, filled_id, encrypted.ciphertext, encrypted.nonce, encrypted.key_version),
        )
        conn.commit()
        before = conn.execute("SELECT value_ciphertext, value_nonce, value_key_version FROM field_values WHERE field_id=? AND account_id=?", (field_id, filled_id)).fetchone()
        before = (bytes(before["value_ciphertext"]), bytes(before["value_nonce"]), before["value_key_version"])

    response = client.post(
        "/accounts/bulk/field/add",
        data={"service_id": service_id, "account_ids": account_ids, "field_name": "Existente"},
    )

    assert response.status_code == 302
    assert "ok=bulk_field_created" in response.headers["Location"]
    with app.app_context():
        conn = get_db()
        from service_manager.crypto import decrypt_secret, EncryptedValue
        assert conn.execute("SELECT COUNT(*) FROM custom_fields WHERE service_id=? AND name='Existente'", (service_id,)).fetchone()[0] == 1
        after = conn.execute("SELECT value_ciphertext, value_nonce, value_key_version FROM field_values WHERE field_id=? AND account_id=?", (field_id, filled_id)).fetchone()
        assert (bytes(after["value_ciphertext"]), bytes(after["value_nonce"]), after["value_key_version"]) == before
        assert decrypt_secret(EncryptedValue(after["value_ciphertext"], after["value_nonce"], after["value_key_version"]), aad=account_field_aad(filled_id, field_id)) == "mantido"
        empty = conn.execute("SELECT value_ciphertext, value_nonce, value_key_version FROM field_values WHERE field_id=? AND account_id=?", (field_id, empty_id)).fetchone()
        assert decrypt_secret(EncryptedValue(empty["value_ciphertext"], empty["value_nonce"], empty["value_key_version"]), aad=account_field_aad(empty_id, field_id)) == ""
        event = conn.execute("SELECT metadata_json FROM audit_events WHERE action='accounts.bulk_field_created' ORDER BY id DESC LIMIT 1").fetchone()
        metadata = json.loads(event["metadata_json"])
        assert metadata == {"count": 2, "created_count": 1, "field_id": field_id}


def test_bulk_field_add_dedupes_selection(app, client):
    login_admin(client)
    service_id, account_ids, _ = seed_bulk_data(app)

    response = client.post(
        "/accounts/bulk/field/add",
        data={"service_id": service_id, "account_ids": account_ids + account_ids, "field_name": "Dup"},
    )

    assert response.status_code == 302
    with app.app_context():
        conn = get_db()
        field_id = conn.execute("SELECT id FROM custom_fields WHERE service_id=? AND name='Dup'", (service_id,)).fetchone()["id"]
        assert conn.execute("SELECT COUNT(*) FROM field_values WHERE field_id=?", (field_id,)).fetchone()[0] == 2
        event = conn.execute("SELECT metadata_json FROM audit_events WHERE action='accounts.bulk_field_created' ORDER BY id DESC LIMIT 1").fetchone()
        assert json.loads(event["metadata_json"]) == {"count": 2, "created_count": 2, "field_id": field_id}


def test_bulk_field_add_rejects_invalid_requests(app, client):
    login_admin(client)
    service_id, account_ids, foreign_id = seed_bulk_data(app)

    foreign = client.post("/accounts/bulk/field/add", data={"service_id": service_id, "account_ids": [foreign_id], "field_name": "N"})
    empty_sel = client.post("/accounts/bulk/field/add", data={"service_id": service_id, "field_name": "N"})
    over_limit = client.post("/accounts/bulk/field/add", data={"service_id": service_id, "account_ids": list(range(1, 202)), "field_name": "N"})
    empty_name = client.post("/accounts/bulk/field/add", data={"service_id": service_id, "account_ids": account_ids, "field_name": "   "}, headers={"Accept": "application/json"})
    long_name = client.post("/accounts/bulk/field/add", data={"service_id": service_id, "account_ids": account_ids, "field_name": "x" * 101})

    assert foreign.status_code == 404
    assert empty_sel.status_code == 400
    assert over_limit.status_code == 400
    assert empty_name.status_code == 400
    assert empty_name.get_json() == {"error": "Campo inválido"}
    assert long_name.status_code == 400
    with app.app_context():
        conn = get_db()
        assert conn.execute("SELECT COUNT(*) FROM custom_fields WHERE service_id=?", (service_id,)).fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM field_values").fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM audit_events WHERE action='accounts.bulk_field_created'").fetchone()[0] == 0


def test_bulk_field_add_requires_editor_role(app, client):
    service_id, account_ids, _ = seed_bulk_data(app)
    login_operator(app, client)
    with app.app_context():
        conn = get_db()
        user_id = conn.execute("SELECT id FROM users WHERE username='operator'").fetchone()["id"]
        conn.execute(
            "INSERT INTO service_members (user_id, service_id, role, created_at) VALUES (?, ?, 'viewer', '2026-01-01T00:00:00+00:00')",
            (user_id, service_id),
        )
        conn.commit()

    response = client.post(
        "/accounts/bulk/field/add",
        data={"service_id": service_id, "account_ids": account_ids, "field_name": "Bloqueado"},
    )

    assert response.status_code == 403
    with app.app_context():
        conn = get_db()
        assert conn.execute("SELECT COUNT(*) FROM custom_fields WHERE service_id=?", (service_id,)).fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM field_values").fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM audit_events WHERE action='accounts.bulk_field_created'").fetchone()[0] == 0

def test_audit_view_filters_exports_and_requires_admin(app, client):
    login_admin(client)
    with client.session_transaction() as session:
        session["reauthenticated_at"] = time.time()
    service_id, account_ids, _ = seed_bulk_data(app)
    client.post("/accounts/bulk/status", data={"service_id": service_id, "account_ids": account_ids, "status": "ativo"})

    response = client.get("/admin/audit")
    assert response.status_code == 200
    body = response.get_data(as_text=True)
    assert "accounts.bulk_status" in body
    assert "Cadeia íntegra" in body

    filtered = client.get("/admin/audit?action=accounts.bulk_status").get_data(as_text=True)
    assert "login.succeeded" not in filtered
    assert "accounts.bulk_status" in filtered

    exported = client.get("/admin/audit.csv?action=accounts.bulk_status")
    assert exported.status_code == 200
    assert exported.get_data(as_text=True).startswith("\ufeffid,occurred_at,usuario,action,target_type,target_id,metadata_json,source_ip")

    client.post("/logout")
    login_operator(app, client)
    assert client.get("/admin/audit").status_code == 403
    assert client.get("/admin/audit.csv").status_code == 403


def test_audit_requires_recent_reauth_and_ip_hash_controls(app, client):
    login_admin(client)
    service_id, account_ids, _ = seed_bulk_data(app)
    client.post("/accounts/bulk/status", data={"service_id": service_id, "account_ids": account_ids, "status": "ativo"})

    # No recent reauth -> both view and export are 403.
    with client.session_transaction() as session:
        session["reauthenticated_at"] = None
    assert client.get("/admin/audit").status_code == 403
    assert client.get("/admin/audit.csv").status_code == 403

    # Real reauth flow restores access: POST /reauth returns 204, then view succeeds.
    reauth = client.post("/reauth", data={"password": ADMIN_PASSWORD})
    assert reauth.status_code == 204
    assert client.get("/admin/audit").status_code == 200

    # Literal partial IP filter: matching value keeps rows, non-matching empties them.
    matched = client.get("/admin/audit?source_ip=127").get_data(as_text=True)
    assert "accounts.bulk_status" in matched
    missing = client.get("/admin/audit?source_ip=203.0.113.9").get_data(as_text=True)
    assert "accounts.bulk_status" not in missing

    # LIKE wildcard input is treated literally (no injection): '%' matches nothing literal.
    wildcard = client.get("/admin/audit?source_ip=%25").get_data(as_text=True)
    assert "accounts.bulk_status" not in wildcard

    # Filter survives on the export link in rendered HTML.
    page = client.get("/admin/audit?source_ip=127").get_data(as_text=True)
    assert "source_ip=127" in page

    # IP filter applies identically to CSV export: match keeps rows, miss/literal-wildcard empty them.
    csv_match = client.get("/admin/audit.csv?source_ip=127").get_data(as_text=True)
    assert "accounts.bulk_status" in csv_match
    csv_miss = client.get("/admin/audit.csv?source_ip=203.0.113.9").get_data(as_text=True)
    assert "accounts.bulk_status" not in csv_miss
    csv_wildcard = client.get("/admin/audit.csv?source_ip=%25").get_data(as_text=True)
    assert "accounts.bulk_status" not in csv_wildcard

    # CSV includes hash columns as lowercase 64-char hex matching the DB rows.
    exported = client.get("/admin/audit.csv").get_data(as_text=True)
    header = exported.splitlines()[0]
    assert header.endswith("source_ip,previous_hash,event_hash")
    with app.app_context():
        row = get_db().execute("SELECT previous_hash, event_hash FROM audit_events ORDER BY id DESC LIMIT 1").fetchone()
    prev_hex = row["previous_hash"].hex()
    event_hex = row["event_hash"].hex()
    assert prev_hex == prev_hex.lower() and len(prev_hex) == 64
    assert len(event_hex) == 64
    assert prev_hex in exported and event_hex in exported

    # Operator is denied even with a recent reauth (role gate precedes reauth window).
    client.post("/logout")
    login_operator(app, client)
    with client.session_transaction() as session:
        session["reauthenticated_at"] = time.time()
    assert client.get("/admin/audit").status_code == 403
    assert client.get("/admin/audit.csv").status_code == 403


def test_audit_pagination_links_preserve_ip_filter(app, client):
    login_admin(client)
    with client.session_transaction() as session:
        session["reauthenticated_at"] = time.time()
    # Seed >50 events sharing a source IP so both pagination directions render.
    with app.app_context():
        conn = get_db()
        with transaction(conn):
            for i in range(120):
                append_audit_event(
                    conn,
                    action="probe.seeded",
                    target_type="probe",
                    target_id=i,
                    source_ip="198.51.100.7",
                )

    page2 = client.get("/admin/audit?source_ip=198.51.100.7&page=2").get_data(as_text=True)
    # The pagination anchors themselves must carry both page target and the IP filter,
    # not merely the form/export links.
    prev_href = re.search(r'href="([^"]*page=1[^"]*)"[^>]*>Anterior', page2)
    next_href = re.search(r'href="([^"]*page=3[^"]*)"[^>]*>Próxima', page2)
    assert prev_href and "source_ip=198.51.100.7" in prev_href.group(1)
    assert next_href and "source_ip=198.51.100.7" in next_href.group(1)
    assert "probe.seeded" in page2


def test_coverage_matrix_renders_for_authenticated_user(app, client):
    service_id, account_ids, _ = seed_bulk_data(app)
    login_operator(app, client)
    with app.app_context():
        conn = get_db()
        user_id = conn.execute("SELECT id FROM users WHERE username='operator'").fetchone()["id"]
        conn.execute(
            "INSERT INTO service_members (user_id, service_id, role, created_at) VALUES (?, ?, 'viewer', '2026-01-01T00:00:00+00:00')",
            (user_id, service_id),
        )
        conn.commit()

    response = client.get("/coverage")

    assert response.status_code == 200
    body = response.get_data(as_text=True)
    assert "Bulk" in body
    assert "bulk-one@example.test" in body
    assert f"/?service={service_id}#row-{account_ids[0]}" in body
    assert 'id="coverage-filter"' in body
    # Missing-registration data attributes are emitted for every visible service.
    assert f'data-reg-svc-{service_id}="0"' in body
    assert "coverage-service-filter" in body
    assert 'data-coverage-service' in body
    assert "missing-registration" in body


def test_coverage_emits_full_case_matrix_data(app, client):
    login_admin(client)
    with app.app_context():
        conn = get_db()
        alpha = inserted_id(conn.execute("INSERT INTO services (name) VALUES ('Alpha')"))
        beta = inserted_id(conn.execute("INSERT INTO services (name) VALUES ('Beta')"))

        def make_account(email: str) -> int:
            return inserted_id(conn.execute(
                "INSERT INTO accounts (email, password_ciphertext, password_nonce, password_key_version) VALUES (?, ?, ?, 1)",
                (email, b"x", b"0" * 12),
            ))

        def link(account_id: int, service_id: int, status: str, registered: int) -> None:
            conn.execute(
                "INSERT INTO account_service (account_id, service_id, status, registered) VALUES (?, ?, ?, ?)",
                (account_id, service_id, status, registered),
            )

        # all-registered: registered in both services.
        full = make_account("full@example.test")
        link(full, alpha, "ativo", 1)
        link(full, beta, "ativo", 1)
        # one-gap: registered in alpha, linked-but-unregistered in beta.
        gap = make_account("gap@example.test")
        link(gap, alpha, "ativo", 1)
        link(gap, beta, "nunca", 0)
        # no-link: linked only to alpha (beta cell must be synthesized as 0).
        nolink = make_account("nolink@example.test")
        link(nolink, alpha, "ativo", 1)
        # no-registration: linked to both but registered in neither.
        none_reg = make_account("nonereg@example.test")
        link(none_reg, alpha, "nunca", 0)
        link(none_reg, beta, "inativo", 0)
        # multi-active: active in both services.
        multi = make_account("multi@example.test")
        link(multi, alpha, "ativo", 1)
        link(multi, beta, "ativo", 1)
        conn.commit()

    body = client.get("/coverage").get_data(as_text=True)
    # Extract each row's opening tag for attribute assertions.
    def row_tag(email: str) -> str:
        match = re.search(rf'(<tr data-coverage-row[^>]*>)(?:(?!</tr>).)*?{re.escape(email)}', body, re.S)
        assert match, f"row for {email} not found"
        return match.group(1)

    full_tag = row_tag("full@example.test")
    assert f'data-reg-svc-{alpha}="1"' in full_tag and f'data-reg-svc-{beta}="1"' in full_tag
    assert 'data-reg-count="2"' in full_tag and 'data-active-count="2"' in full_tag

    gap_tag = row_tag("gap@example.test")
    assert f'data-reg-svc-{alpha}="1"' in gap_tag and f'data-reg-svc-{beta}="0"' in gap_tag

    nolink_tag = row_tag("nolink@example.test")
    # Beta has no account_service row -> synthesized as 0.
    assert f'data-reg-svc-{alpha}="1"' in nolink_tag and f'data-reg-svc-{beta}="0"' in nolink_tag

    nonereg_tag = row_tag("nonereg@example.test")
    assert f'data-reg-svc-{alpha}="0"' in nonereg_tag and f'data-reg-svc-{beta}="0"' in nonereg_tag
    assert 'data-reg-count="0"' in nonereg_tag

    multi_tag = row_tag("multi@example.test")
    assert 'data-active-count="2"' in multi_tag


def test_coverage_matrix_requires_authentication(client):
    response = client.get("/coverage")

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/login")


def _seed_rotation_account(app, *, email="rot@example.test", password_changed_at=None, service_days=None, link_days=None, due_at=None):
    with app.app_context():
        conn = get_db()
        conn.execute("INSERT OR REPLACE INTO app_settings (key, value) VALUES ('rotation_enabled', '1')")
        service_id = inserted_id(conn.execute("INSERT INTO services (name) VALUES ('Rotate')"))
        if service_days is not None:
            conn.execute("UPDATE services SET rotation_days=? WHERE id=?", (service_days, service_id))
        account_id = inserted_id(conn.execute(
            "INSERT INTO accounts (email, password_ciphertext, password_nonce, password_key_version, password_changed_at) VALUES (?, ?, ?, 1, ?)",
            (email, b"x", b"0" * 12, password_changed_at),
        ))
        conn.execute(
            "INSERT INTO account_service (account_id, service_id, status, registered, rotation_days, rotation_due_at) VALUES (?, ?, 'ativo', 1, ?, ?)",
            (account_id, service_id, link_days, due_at),
        )
        conn.commit()
    return service_id, account_id


def test_service_rotation_policy_updates_and_audits(app, client):
    login_admin(client)
    service_id, _ = _seed_rotation_account(app)

    resp = client.post(f"/service/{service_id}/rotation-policy", data={"rotation_days": "45"})
    assert resp.status_code == 302
    with app.app_context():
        conn = get_db()
        assert conn.execute("SELECT rotation_days FROM services WHERE id=?", (service_id,)).fetchone()[0] == 45
        meta = json.loads(conn.execute("SELECT metadata_json FROM audit_events WHERE action='rotation.policy_updated' ORDER BY id DESC LIMIT 1").fetchone()[0])
        assert meta == {"service_id": service_id, "rotation_days": 45, "rotation_due_at": None}

    # Blank clears the policy.
    assert client.post(f"/service/{service_id}/rotation-policy", data={"rotation_days": ""}).status_code == 302
    with app.app_context():
        assert get_db().execute("SELECT rotation_days FROM services WHERE id=?", (service_id,)).fetchone()[0] is None

    # Invalid interval rejected.
    assert client.post(f"/service/{service_id}/rotation-policy", data={"rotation_days": "abc"}).status_code == 400
    assert client.post(f"/service/{service_id}/rotation-policy", data={"rotation_days": "0"}).status_code == 400
    # Nonexistent service -> 404.
    assert client.post("/service/999999/rotation-policy", data={"rotation_days": "10"}).status_code == 404


def test_account_rotation_policy_updates_selected_link(app, client):
    login_admin(client)
    service_id, account_id = _seed_rotation_account(app)

    resp = client.post(f"/accounts/{account_id}/rotation-policy", data={"service_id": service_id, "rotation_days": "15", "rotation_due_at": "2026-08-01"})
    assert resp.status_code == 302
    with app.app_context():
        conn = get_db()
        row = conn.execute("SELECT rotation_days, rotation_due_at FROM account_service WHERE account_id=? AND service_id=?", (account_id, service_id)).fetchone()
        assert row["rotation_days"] == 15 and row["rotation_due_at"] == "2026-08-01"
        meta = json.loads(conn.execute("SELECT metadata_json FROM audit_events WHERE action='rotation.policy_updated' ORDER BY id DESC LIMIT 1").fetchone()[0])
        assert meta == {"service_id": service_id, "rotation_days": 15, "rotation_due_at": "2026-08-01"}

    # Blanks clear both to inherit.
    assert client.post(f"/accounts/{account_id}/rotation-policy", data={"service_id": service_id, "rotation_days": "", "rotation_due_at": ""}).status_code == 302
    with app.app_context():
        row = get_db().execute("SELECT rotation_days, rotation_due_at FROM account_service WHERE account_id=? AND service_id=?", (account_id, service_id)).fetchone()
        assert row["rotation_days"] is None and row["rotation_due_at"] is None

    # Invalid date rejected.
    assert client.post(f"/accounts/{account_id}/rotation-policy", data={"service_id": service_id, "rotation_days": "", "rotation_due_at": "2026/08/01"}).status_code == 400


def test_complete_rotation_replaces_password_and_restarts_schedules(app, client):
    login_admin(client)
    service_id, account_id = _seed_rotation_account(app, password_changed_at="2020-01-01T00:00:00+00:00", link_days=30, due_at="2026-08-01")
    with app.app_context():
        conn = get_db()
        # Add a second linked service with its own explicit due override.
        other_id = inserted_id(conn.execute("INSERT INTO services (name) VALUES ('Second')"))
        conn.execute("INSERT INTO account_service (account_id, service_id, status, registered, rotation_due_at) VALUES (?, ?, 'ativo', 1, '2026-09-01')", (account_id, other_id))
        conn.commit()
        before = conn.execute("SELECT password_ciphertext, password_nonce, password_changed_at FROM accounts WHERE id=?", (account_id,)).fetchone()

    resp = client.post(f"/accounts/{account_id}/rotation", data={"service_id": service_id, "outcome": "completed", "new_password": "brand-new-rotation-secret"})
    assert resp.status_code == 302
    assert "ok=rotation_completed" in resp.headers["Location"]
    with app.app_context():
        conn = get_db()
        after = conn.execute("SELECT password_ciphertext, password_nonce, password_changed_at FROM accounts WHERE id=?", (account_id,)).fetchone()
        # Encrypted envelope changed and timestamp advanced.
        assert after["password_ciphertext"] != before["password_ciphertext"]
        assert after["password_changed_at"] != before["password_changed_at"]
        # Every explicit due override cleared so each service restarts from its interval.
        overrides = conn.execute("SELECT COUNT(*) FROM account_service WHERE account_id=? AND rotation_due_at IS NOT NULL", (account_id,)).fetchone()[0]
        assert overrides == 0
        # New plaintext never appears in audit metadata.
        audit_rows = conn.execute("SELECT metadata_json FROM audit_events WHERE action='rotation.completed'").fetchall()
        assert audit_rows and all("brand-new-rotation-secret" not in (r[0] or "") for r in audit_rows)


def test_incomplete_rotation_changes_nothing_and_audits(app, client):
    login_admin(client)
    service_id, account_id = _seed_rotation_account(app, password_changed_at="2020-01-01T00:00:00+00:00", link_days=30, due_at="2026-08-01")
    with app.app_context():
        conn = get_db()
        before = conn.execute("SELECT password_ciphertext, password_changed_at FROM accounts WHERE id=?", (account_id,)).fetchone()
        due_before = conn.execute("SELECT rotation_due_at FROM account_service WHERE account_id=? AND service_id=?", (account_id, service_id)).fetchone()[0]

    resp = client.post(f"/accounts/{account_id}/rotation", data={"service_id": service_id, "outcome": "incomplete"})
    assert resp.status_code == 302
    assert "ok=rotation_incomplete" in resp.headers["Location"]
    with app.app_context():
        conn = get_db()
        after = conn.execute("SELECT password_ciphertext, password_changed_at FROM accounts WHERE id=?", (account_id,)).fetchone()
        assert after["password_ciphertext"] == before["password_ciphertext"]
        assert after["password_changed_at"] == before["password_changed_at"]
        due_after = conn.execute("SELECT rotation_due_at FROM account_service WHERE account_id=? AND service_id=?", (account_id, service_id)).fetchone()[0]
        assert due_after == due_before
        assert conn.execute("SELECT COUNT(*) FROM audit_events WHERE action='rotation.incomplete_marked'").fetchone()[0] == 1


def test_complete_rotation_requires_password_for_completed(app, client):
    login_admin(client)
    service_id, account_id = _seed_rotation_account(app)
    # completed without a password is rejected; no vault-skip path.
    assert client.post(f"/accounts/{account_id}/rotation", data={"service_id": service_id, "outcome": "completed", "new_password": ""}).status_code == 400
    # invalid outcome rejected.
    assert client.post(f"/accounts/{account_id}/rotation", data={"service_id": service_id, "outcome": "bogus"}).status_code == 400


def test_rotation_view_lists_due_and_enforces_role(app, client):
    login_admin(client)
    service_id, account_id = _seed_rotation_account(app, password_changed_at="2020-01-01T00:00:00+00:00", service_days=30)

    resp = client.get(f"/rotation?service={service_id}")
    assert resp.status_code == 200
    body = resp.get_data(as_text=True)
    assert "rot@example.test" in body
    assert 'data-rotation="overdue"' in body

    # Operator without membership is denied.
    client.post("/logout")
    login_operator(app, client)
    assert client.get(f"/rotation?service={service_id}").status_code == 403


def test_index_shows_rotation_column_counts_and_filter(app, client):
    login_admin(client)
    service_id, account_id = _seed_rotation_account(app, password_changed_at="2020-01-01T00:00:00+00:00", service_days=30)

    body = client.get(f"/?service={service_id}").get_data(as_text=True)
    assert 'id="filter-rotation"' in body
    rot_start = body.index('data-column-filter="rotation"')
    rot_th = body[rot_start:body.index("</th>", rot_start)]
    assert 'id="filter-rotation"' in rot_th
    assert 'data-rotation="overdue"' in body
    assert "data-rotation-overdue-count" in body
    assert 'href="/rotation?service=' in body


def test_update_password_clears_due_override_and_audits(app, client):
    login_admin(client)
    service_id, account_id = _seed_rotation_account(app, password_changed_at="2020-01-01T00:00:00+00:00", link_days=30, due_at="2026-08-01")

    # A real password change through update() clears the explicit due override.
    resp = client.post(f"/accounts/{account_id}", data={"service_id": service_id, "email": "rot@example.test", "password": "updated-rotation-secret"})
    assert resp.status_code == 302
    with app.app_context():
        conn = get_db()
        override = conn.execute("SELECT rotation_due_at FROM account_service WHERE account_id=? AND service_id=?", (account_id, service_id)).fetchone()[0]
        assert override is None
        meta = json.loads(conn.execute("SELECT metadata_json FROM audit_events WHERE action='account.updated' ORDER BY id DESC LIMIT 1").fetchone()[0])
        assert meta["password_changed"] is True

    # Email-only update preserves rotation state (re-seed an override first).
    with app.app_context():
        conn = get_db()
        conn.execute("UPDATE account_service SET rotation_due_at='2026-08-01' WHERE account_id=? AND service_id=?", (account_id, service_id))
        conn.commit()
    resp = client.post(f"/accounts/{account_id}", data={"service_id": service_id, "email": "renamed-rot@example.test", "password": ""})
    assert resp.status_code == 302
    with app.app_context():
        override = get_db().execute("SELECT rotation_due_at FROM account_service WHERE account_id=? AND service_id=?", (account_id, service_id)).fetchone()[0]
        assert override == "2026-08-01"


def test_add_and_import_set_password_changed_at(app, client):
    login_admin(client)
    with app.app_context():
        service_id = inserted_id(get_db().execute("INSERT INTO services (name) VALUES ('Onboard')"))
        get_db().commit()

    # add() sets password_changed_at on creation.
    resp = client.post("/add", data={"service_id": service_id, "email": "created@example.test", "password": "created-secret-value", "status": "ativo"})
    assert resp.status_code == 302
    with app.app_context():
        ts = get_db().execute("SELECT password_changed_at FROM accounts WHERE email='created@example.test'").fetchone()[0]
        assert ts is not None

    # import_bulk() sets password_changed_at on each inserted row.
    csv_bytes = b"email,password,status\nimported@example.test,imported-secret,ativo\n"
    resp = client.post(
        "/import",
        data={"service_id": str(service_id), "file": (io.BytesIO(csv_bytes), "accounts.csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 302
    with app.app_context():
        ts = get_db().execute("SELECT password_changed_at FROM accounts WHERE email='imported@example.test'").fetchone()[0]
        assert ts is not None


def _seed_link(app, *, service_name="Frozen", email="frozen@example.test", changed_at="2020-01-01T00:00:00+00:00", link_days=30, due_at: str | None = "2026-08-01"): 
    """Seed a service+account+link WITHOUT enabling rotation (feature off by default)."""
    with app.app_context():
        conn = get_db()
        service_id = inserted_id(conn.execute("INSERT INTO services (name) VALUES (?)", (service_name,)))
        account_id = inserted_id(conn.execute(
            "INSERT INTO accounts (email, password_ciphertext, password_nonce, password_key_version, password_changed_at) VALUES (?, ?, ?, 1, ?)",
            (email, b"x", b"0" * 12, changed_at),
        ))
        conn.execute(
            "INSERT INTO account_service (account_id, service_id, status, registered, rotation_days, rotation_due_at) VALUES (?, ?, 'ativo', 1, ?, ?)",
            (account_id, service_id, link_days, due_at),
        )
        conn.commit()
    return service_id, account_id


def test_rotation_disabled_by_default_gates_every_endpoint(app, client):
    login_admin(client)
    service_id, account_id = _seed_link(app)
    # All four rotation endpoints return 404 while the feature is globally disabled.
    assert client.get(f"/rotation?service={service_id}").status_code == 404
    assert client.post(f"/service/{service_id}/rotation-policy", data={"rotation_days": "45"}).status_code == 404
    assert client.post(f"/accounts/{account_id}/rotation-policy", data={"service_id": service_id, "rotation_days": "15"}).status_code == 404
    assert client.post(f"/accounts/{account_id}/rotation", data={"service_id": service_id, "outcome": "completed", "new_password": "x-very-secret-value"}).status_code == 404
    # The gate precedes payload validation: even malformed requests 404, not 400.
    assert client.post(f"/accounts/{account_id}/rotation", data={"service_id": service_id, "outcome": "bogus"}).status_code == 404
    assert client.post("/service/999999/rotation-policy", data={"rotation_days": "10"}).status_code == 404


def test_disabled_rotation_posts_mutate_nothing_and_leave_no_audit(app, client):
    login_admin(client)
    service_id, account_id = _seed_link(app)
    with app.app_context():
        conn = get_db()
        before = conn.execute("SELECT password_ciphertext, password_changed_at FROM accounts WHERE id=?", (account_id,)).fetchone()
        svc_before = conn.execute("SELECT rotation_days FROM services WHERE id=?", (service_id,)).fetchone()[0]
        link_before = conn.execute("SELECT rotation_days, rotation_due_at FROM account_service WHERE account_id=? AND service_id=?", (account_id, service_id)).fetchone()
        audit_before = conn.execute("SELECT COUNT(*) FROM audit_events").fetchone()[0]

    client.post(f"/service/{service_id}/rotation-policy", data={"rotation_days": "45"})
    client.post(f"/accounts/{account_id}/rotation-policy", data={"service_id": service_id, "rotation_days": "15", "rotation_due_at": "2026-09-09"})
    client.post(f"/accounts/{account_id}/rotation", data={"service_id": service_id, "outcome": "completed", "new_password": "brand-new-secret-value"})

    with app.app_context():
        conn = get_db()
        after = conn.execute("SELECT password_ciphertext, password_changed_at FROM accounts WHERE id=?", (account_id,)).fetchone()
        assert after["password_ciphertext"] == before["password_ciphertext"]
        assert after["password_changed_at"] == before["password_changed_at"]
        assert conn.execute("SELECT rotation_days FROM services WHERE id=?", (service_id,)).fetchone()[0] == svc_before
        link_after = conn.execute("SELECT rotation_days, rotation_due_at FROM account_service WHERE account_id=? AND service_id=?", (account_id, service_id)).fetchone()
        assert (link_after["rotation_days"], link_after["rotation_due_at"]) == (link_before["rotation_days"], link_before["rotation_due_at"])
        assert conn.execute("SELECT COUNT(*) FROM audit_events").fetchone()[0] == audit_before
        assert conn.execute("SELECT COUNT(*) FROM audit_events WHERE action LIKE 'rotation.%'").fetchone()[0] == 0


def test_index_hides_rotation_column_and_filter_when_disabled(app, client):
    login_admin(client)
    service_id, _ = _seed_link(app)
    body = client.get(f"/?service={service_id}").get_data(as_text=True)
    assert 'id="filter-rotation"' not in body
    assert 'data-column-filter="rotation"' not in body
    assert "cell-rotation" not in body
    assert "data-rotation=" not in body
    assert "Ver rotação" not in body
    assert body.count("<th scope") == 6


def test_add_and_import_record_password_changed_at_regardless_of_flag(app, client):
    # password_changed_at is factual data, always recorded; the flag only gates the feature surface.
    login_admin(client)
    with app.app_context():
        service_id = inserted_id(get_db().execute("INSERT INTO services (name) VALUES ('OnboardOff')"))
        get_db().commit()
    assert client.post("/add", data={"service_id": service_id, "email": "off-created@example.test", "password": "created-secret-value", "status": "ativo"}).status_code == 302
    csv_bytes = b"email,password,status\noff-imported@example.test,imported-secret,ativo\n"
    assert client.post("/import", data={"service_id": str(service_id), "file": (io.BytesIO(csv_bytes), "accounts.csv")}, content_type="multipart/form-data").status_code == 302
    with app.app_context():
        conn = get_db()
        assert conn.execute("SELECT password_changed_at FROM accounts WHERE email='off-created@example.test'").fetchone()[0] is not None
        assert conn.execute("SELECT password_changed_at FROM accounts WHERE email='off-imported@example.test'").fetchone()[0] is not None


def test_update_password_records_fresh_timestamp_even_when_disabled(app, client):
    # A real password change always advances password_changed_at and clears due overrides,
    # so re-enabling never inherits a stale "last changed" date for a freshly rotated secret.
    login_admin(client)
    service_id, account_id = _seed_link(app, email="freeze-upd@example.test")
    with app.app_context():
        before = get_db().execute("SELECT password_ciphertext, password_changed_at FROM accounts WHERE id=?", (account_id,)).fetchone()
    resp = client.post(f"/accounts/{account_id}", data={"service_id": service_id, "email": "freeze-upd@example.test", "password": "rotated-off-secret-value"})
    assert resp.status_code == 302
    with app.app_context():
        conn = get_db()
        after = conn.execute("SELECT password_ciphertext, password_changed_at FROM accounts WHERE id=?", (account_id,)).fetchone()
        assert after["password_ciphertext"] != before["password_ciphertext"]
        assert after["password_changed_at"] != before["password_changed_at"]
        assert after["password_changed_at"] is not None
        due = conn.execute("SELECT rotation_due_at FROM account_service WHERE account_id=? AND service_id=?", (account_id, service_id)).fetchone()[0]
        assert due is None
        meta = json.loads(conn.execute("SELECT metadata_json FROM audit_events WHERE action='account.updated' ORDER BY id DESC LIMIT 1").fetchone()[0])
        assert meta["password_changed"] is True


def test_enabling_rotation_resumes_from_preexisting_metadata(app, client):
    login_admin(client)
    # Pre-existing account with real history/policy captured before the feature was toggled.
    service_id, account_id = _seed_link(app, email="resume@example.test", changed_at="2020-01-01T00:00:00+00:00", link_days=30, due_at=None)
    enable_rotation(app)
    # The overdue state is computed from the preserved timestamp+policy, not reset to unknown.
    body = client.get(f"/rotation?service={service_id}").get_data(as_text=True)
    assert "resume@example.test" in body
    assert 'data-rotation="overdue"' in body


def test_settings_view_reflects_state_and_requires_admin(app, client):
    login_admin(client)
    body = client.get("/admin/settings").get_data(as_text=True)
    assert "Rotação de credenciais" in body
    assert 'name="rotation_enabled"' in body
    # No checkbox checked while disabled.
    assert "checked" not in body.split('name="rotation_enabled"')[1].split(">")[0]
    # Operator is denied the admin settings page.
    client.post("/logout")
    login_operator(app, client)
    assert client.get("/admin/settings").status_code == 403


def test_settings_update_requires_reauth_then_toggles(app, client):
    login_admin(client)
    # Without a recent reauth the POST is refused and nothing is persisted.
    assert client.post("/admin/settings", data={"rotation_enabled": "1"}).status_code == 403
    with app.app_context():
        assert get_db().execute("SELECT COUNT(*) FROM app_settings WHERE key='rotation_enabled'").fetchone()[0] == 0
    # With a recent reauth the toggle persists and audits.
    with client.session_transaction() as session:
        session["reauthenticated_at"] = time.time()
    resp = client.post("/admin/settings", data={"rotation_enabled": "1"})
    assert resp.status_code == 302 and "ok=settings_updated" in resp.headers["Location"]
    with app.app_context():
        conn = get_db()
        assert conn.execute("SELECT value FROM app_settings WHERE key='rotation_enabled'").fetchone()[0] == "1"
        meta = json.loads(conn.execute("SELECT metadata_json FROM audit_events WHERE action='settings.rotation_enabled_updated' ORDER BY id DESC LIMIT 1").fetchone()[0])
        assert meta == {"enabled": True}
    # Now the rotation endpoints are reachable.
    with app.app_context():
        sid = inserted_id(get_db().execute("INSERT INTO services (name) VALUES ('AfterToggle')"))
        get_db().commit()
    assert client.get(f"/rotation?service={sid}").status_code == 200
    # Unchecking disables again.
    with client.session_transaction() as session:
        session["reauthenticated_at"] = time.time()
    assert client.post("/admin/settings", data={}).status_code == 302
    with app.app_context():
        assert get_db().execute("SELECT value FROM app_settings WHERE key='rotation_enabled'").fetchone()[0] == "0"
    assert client.get(f"/rotation?service={sid}").status_code == 404


def test_settings_operator_post_is_forbidden(app, client):
    login_operator(app, client)
    with client.session_transaction() as session:
        session["reauthenticated_at"] = time.time()
    assert client.post("/admin/settings", data={"rotation_enabled": "1"}).status_code == 403
    with app.app_context():
        assert get_db().execute("SELECT COUNT(*) FROM app_settings").fetchone()[0] == 0
