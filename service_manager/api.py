from __future__ import annotations

import csv
import hmac
import io
import json
import re
import sqlite3
import tempfile
import unicodedata
from datetime import UTC, date, datetime, timedelta
from typing import Any

from flask import Blueprint, Response, current_app, g, jsonify, request, send_file
from flask.typing import ResponseReturnValue
from itsdangerous import BadSignature, URLSafeSerializer
from werkzeug.exceptions import HTTPException

from service_manager.audit import verify_audit_chain
from service_manager.auth import source_ip
from service_manager.authorization import (
    accessible_services,
    require_account_role,
    require_accounts_role,
    require_service_role,
)
from service_manager.crypto import EncryptedValue, account_field_aad, account_password_aad, decrypt_secret
from service_manager.db import get_db, transaction
from service_manager.imports import ImportFormatError, has_allowed_upload_mimetype, parse_import_file
from service_manager.operations import (
    DomainError,
    NotFoundError,
    bulk_create_field,
    bulk_delete_accounts,
    bulk_set_field,
    bulk_update_registered,
    bulk_update_status,
    change_user_active,
    change_user_role,
    complete_rotation,
    create_account,
    create_api_key,
    create_service,
    create_user,
    create_webhook,
    delete_webhook,
    test_webhook,
    update_webhook,
    delete_account,
    delete_field_value,
    delete_service,
    grant_membership,
    hash_api_key_raw,
    import_accounts,
    list_api_keys,
    normalize_status,
    parse_api_key_token,
    principal_from_api_key,
    reveal_account_password,
    revoke_api_key,
    revoke_membership,
    set_account_rotation_policy,
    set_rotation_enabled_setting,
    set_service_rotation_policy,
    touch_api_key_last_used,
    update_account,
    update_account_registered,
    update_account_status,
    update_field_value,
    update_service_preferences,
    upsert_field_values,
    valid_email,
    valid_name,
    valid_secret,
    audit as domain_audit,
)
from service_manager.webhooks import (
    list_webhook_configs,
    webhook_event_types,
)

api = Blueprint("api", __name__, url_prefix="/api/v1")

_ACCOUNT_PAGE_SIZE = 100
_COVERAGE_PAGE_SIZE = 100
_EXPORT_LIMIT = 10_000
_STATUS_RANK_SQL = "CASE link.status WHEN 'ativo' THEN 0 WHEN 'nunca' THEN 1 WHEN 'inativo' THEN 2 ELSE 1 END"
_ACCOUNT_SELECT = (
    "SELECT a.id AS id, a.email AS email, a.password_changed_at AS password_changed_at, "
    "link.status AS status, link.registered AS registered, "
    "link.rotation_days AS rotation_days, link.rotation_due_at AS rotation_due_at "
    "FROM account_service AS link JOIN accounts AS a ON a.id = link.account_id"
)
_ROTATION_STATE_SQL = """
CASE
  WHEN link.rotation_due_at IS NOT NULL THEN
    CASE
      WHEN date(link.rotation_due_at) IS NULL THEN 'unknown'
      WHEN CAST(julianday(date(link.rotation_due_at)) - julianday(date(:rot_today)) AS INTEGER) < 0 THEN 'overdue'
      WHEN CAST(julianday(date(link.rotation_due_at)) - julianday(date(:rot_today)) AS INTEGER) <= 7 THEN 'due_soon'
      ELSE 'current'
    END
  WHEN a.password_changed_at IS NULL THEN 'unknown'
  WHEN NOT (a.password_changed_at LIKE '%Z' OR a.password_changed_at LIKE '%+__:__' OR a.password_changed_at LIKE '%-__:__') THEN 'unknown'
  WHEN date(a.password_changed_at) IS NULL THEN 'unknown'
  WHEN COALESCE(link.rotation_days, :rot_sdays) IS NULL THEN 'no_policy'
  ELSE
    CASE
      WHEN CAST(julianday(date(a.password_changed_at, '+' || COALESCE(link.rotation_days, :rot_sdays) || ' days')) - julianday(date(:rot_today)) AS INTEGER) < 0 THEN 'overdue'
      WHEN CAST(julianday(date(a.password_changed_at, '+' || COALESCE(link.rotation_days, :rot_sdays) || ' days')) - julianday(date(:rot_today)) AS INTEGER) <= 7 THEN 'due_soon'
      ELSE 'current'
    END
END
"""
TEMPLATE_ROWS = [
    ("email", "password", "status"),
    ("exemplo1@gmail.com", "SenhaSegura1", "nunca"),
    ("exemplo2@gmail.com", "SenhaSegura2", "ativo"),
]


def _json_error(code: str, message: str, status: int) -> Response:
    response = jsonify(error={"code": code, "message": message})
    response.status_code = status
    response.headers["Cache-Control"] = "no-store"
    if status == 401:
        response.headers["WWW-Authenticate"] = 'Bearer realm="service-manager"'
    return response


def _no_store(response: Response) -> Response:
    response.headers["Cache-Control"] = "no-store"
    return response


def _json(data: Any, status: int = 200) -> Response:
    response = jsonify(data)
    response.status_code = status
    return _no_store(response)


def _empty(status: int = 204) -> Response:
    return _no_store(Response(status=status))


def _principal():
    return principal_from_api_key(g.api_key)


def _json_body() -> dict[str, Any]:
    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        raise DomainError("invalid_json", "JSON inválido")
    return payload


def _require_rotation_enabled(conn: sqlite3.Connection) -> None:
    row = conn.execute("SELECT value FROM app_settings WHERE key='rotation_enabled'").fetchone()
    if row is None or row["value"] != "1":
        raise NotFoundError()


def _parse_rotation_days_value(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool) or not isinstance(value, int):
        raise DomainError("invalid_rotation", "Intervalo inválido")
    if not (1 <= value <= 3650):
        raise DomainError("invalid_rotation", "Intervalo inválido")
    return value


def _parse_rotation_due_at_value(value: Any) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str) or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        raise DomainError("invalid_rotation", "Política de rotação inválida")
    try:
        parsed = date.fromisoformat(value)
    except ValueError as error:
        raise DomainError("invalid_rotation", "Política de rotação inválida") from error
    if parsed.isoformat() != value:
        raise DomainError("invalid_rotation", "Política de rotação inválida")
    return value


def _bulk_ids(raw: Any) -> list[int]:
    if not isinstance(raw, list):
        raise DomainError("invalid_selection", "Seleção inválida")
    try:
        account_ids = [int(item) for item in raw]
    except (TypeError, ValueError) as error:
        raise DomainError("invalid_selection", "Seleção inválida") from error
    if any(account_id <= 0 for account_id in account_ids):
        raise DomainError("invalid_selection", "Seleção inválida")
    unique = list(dict.fromkeys(account_ids))
    if not unique or len(unique) > 200:
        raise DomainError("invalid_selection", "Seleção inválida")
    return unique


def _cursor_serializer(salt: str) -> URLSafeSerializer:
    return URLSafeSerializer(current_app.config["SECRET_KEY"], salt=salt)


def _encode_cursor(salt: str, payload: dict[str, object]) -> str:
    return _cursor_serializer(salt).dumps(payload)


def _decode_cursor(salt: str, token: str) -> dict[str, object]:
    try:
        payload = _cursor_serializer(salt).loads(token)
    except BadSignature as error:
        raise DomainError("invalid_cursor", "Cursor inválido") from error
    if not isinstance(payload, dict):
        raise DomainError("invalid_cursor", "Cursor inválido")
    return payload


def _like_escape(value: str) -> str:
    return value.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


def _sanitize_cell(value: str) -> str:
    return "'" + value if value[:1] in ("=", "+", "-", "@") else value


def _safe_filename_slug(name: str) -> str:
    folded = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii").lower()
    slug = re.sub(r"[^a-z0-9]+", "_", folded).strip("_")[:50].strip("_")
    return slug or "servico"


def _rotation_state(
    password_changed_at: str | None,
    account_days: int | None,
    account_due_at: str | None,
    service_days: int | None,
    *,
    today: date,
) -> dict[str, Any]:
    if account_due_at is not None:
        try:
            due = date.fromisoformat(account_due_at)
        except ValueError:
            return {"state": "unknown", "effective_days": None, "due_at": None, "days_remaining": None}
        days_remaining = (due - today).days
        state = "overdue" if days_remaining < 0 else "due_soon" if days_remaining <= 7 else "current"
        return {
            "state": state,
            "effective_days": account_days if account_days is not None else service_days,
            "due_at": due.isoformat(),
            "days_remaining": days_remaining,
        }
    if not password_changed_at:
        return {"state": "unknown", "effective_days": None, "due_at": None, "days_remaining": None}
    if not (password_changed_at.endswith("Z") or re.search(r"[+-]\d{2}:\d{2}$", password_changed_at)):
        return {"state": "unknown", "effective_days": None, "due_at": None, "days_remaining": None}
    try:
        changed = datetime.fromisoformat(password_changed_at.replace("Z", "+00:00")).date()
    except ValueError:
        return {"state": "unknown", "effective_days": None, "due_at": None, "days_remaining": None}
    effective_days = account_days if account_days is not None else service_days
    if effective_days is None:
        return {"state": "no_policy", "effective_days": None, "due_at": None, "days_remaining": None}
    due = changed + timedelta(days=effective_days)
    days_remaining = (due - today).days
    state = "overdue" if days_remaining < 0 else "due_soon" if days_remaining <= 7 else "current"
    return {
        "state": state,
        "effective_days": effective_days,
        "due_at": due.isoformat(),
        "days_remaining": days_remaining,
    }


def _account_filter_sql(*, service_id, q, status, registered, rot_state, service_days, today):
    clauses = ["link.service_id = :sid"]
    params: dict[str, object] = {"sid": service_id, "rot_today": today.isoformat(), "rot_sdays": service_days}
    if q:
        clauses.append("a.email LIKE :q ESCAPE '\\\\'")
        params["q"] = f"%{_like_escape(q)}%"
    if status:
        clauses.append("link.status = :status")
        params["status"] = status
    if registered is not None:
        clauses.append("link.registered = :registered")
        params["registered"] = registered
    if rot_state:
        clauses.append(f"({_ROTATION_STATE_SQL}) = :rot_state")
        params["rot_state"] = rot_state
    return " AND ".join(clauses), params


def _account_order_sql(sort: str, sql_dir: str) -> str:
    if sort == "status":
        return f"{_STATUS_RANK_SQL} {sql_dir}, a.email COLLATE NOCASE {sql_dir}, a.id {sql_dir}"
    return f"a.email COLLATE NOCASE {sql_dir}, a.id {sql_dir}"


def _account_keyset_sql(sort: str, op: str) -> str:
    if sort == "status":
        rank = _STATUS_RANK_SQL
        return (
            f"(({rank} {op} :cursor_rank) OR "
            f"({rank} = :cursor_rank AND a.email COLLATE NOCASE {op} :cursor_email) OR "
            f"({rank} = :cursor_rank AND a.email COLLATE NOCASE = :cursor_email AND a.id {op} :cursor_id))"
        )
    return (
        f"((a.email COLLATE NOCASE {op} :cursor_email) OR "
        f"(a.email COLLATE NOCASE = :cursor_email AND a.id {op} :cursor_id))"
    )




def _export_fields(conn: sqlite3.Connection, service_id: int) -> tuple[tuple[int, str], ...]:
    return tuple(
        (row["id"], row["name"])
        for row in conn.execute(
            "SELECT id, name FROM custom_fields WHERE service_id=? ORDER BY name COLLATE NOCASE, id",
            (service_id,),
        )
    )


def _iter_export_rows(conn: sqlite3.Connection, service_id: int, fields: tuple[tuple[int, str], ...]):
    field_ids = [field_id for field_id, _ in fields]
    accounts = conn.execute(
        """
        SELECT a.id, a.email, a.password_ciphertext, a.password_nonce, a.password_key_version, link.status, link.registered
        FROM account_service AS link
        JOIN accounts AS a ON a.id = link.account_id
        WHERE link.service_id = ?
        ORDER BY a.email COLLATE NOCASE, a.id
        """,
        (service_id,),
    ).fetchall()
    values_by_pair: dict[tuple[int, int], EncryptedValue] = {}
    if field_ids:
        placeholders = ",".join("?" for _ in field_ids)
        for row in conn.execute(
            f"SELECT field_id, account_id, value_ciphertext, value_nonce, value_key_version FROM field_values WHERE field_id IN ({placeholders})",
            field_ids,
        ):
            values_by_pair[(row["field_id"], row["account_id"])] = EncryptedValue(
                row["value_ciphertext"], row["value_nonce"], row["value_key_version"]
            )
    for account in accounts:
        password = decrypt_secret(
            EncryptedValue(account["password_ciphertext"], account["password_nonce"], account["password_key_version"]),
            aad=account_password_aad(account["id"]),
        )
        field_values = []
        for field_id, _ in fields:
            encrypted = values_by_pair.get((field_id, account["id"]))
            if encrypted is None:
                field_values.append("")
            else:
                field_values.append(decrypt_secret(encrypted, aad=account_field_aad(account["id"], field_id)))
        yield (
            _sanitize_cell(account["email"]),
            _sanitize_cell(password),
            account["status"],
            "1" if account["registered"] else "0",
            *(_sanitize_cell(value) for value in field_values),
        )


def _webhook_resolver():
    return current_app.config.get("WEBHOOK_DNS_RESOLVER")


def _actor_label(username: str | None, metadata_json: str | None) -> str:
    if username:
        return username
    if metadata_json:
        try:
            metadata = json.loads(metadata_json)
        except (TypeError, ValueError):
            metadata = None
        if isinstance(metadata, dict):
            name = metadata.get("api_key_name")
            if isinstance(name, str) and name:
                return f"api:{name}"
    return "sistema"


@api.before_app_request
def authenticate_api_key() -> ResponseReturnValue | None:
    if request.blueprint != "api":
        return None
    header = request.headers.get("Authorization")
    if header is None:
        return _json_error("unauthorized", "Autenticação necessária", 401)
    if header.count(" ") != 1:
        return _json_error("unauthorized", "Cabeçalho Authorization inválido", 401)
    scheme, token = header.split(" ", 1)
    if scheme != "Bearer" or not token:
        return _json_error("unauthorized", "Cabeçalho Authorization inválido", 401)
    raw = parse_api_key_token(token)
    if raw is None:
        return _json_error("unauthorized", "API key inválida", 401)
    digest = hash_api_key_raw(raw)
    conn = get_db()
    row = conn.execute(
        "SELECT id, name, created_at, last_used_at, revoked_at, secret_hash FROM api_keys WHERE secret_hash = ?",
        (digest,),
    ).fetchone()
    if row is None or row["revoked_at"] is not None:
        return _json_error("unauthorized", "API key inválida ou revogada", 401)
    if not hmac.compare_digest(bytes(row["secret_hash"]), digest):
        return _json_error("unauthorized", "API key inválida ou revogada", 401)
    g.api_key = {
        "id": row["id"],
        "name": row["name"],
        "created_at": row["created_at"],
        "last_used_at": row["last_used_at"],
    }
    g.current_user = {
        "id": None,
        "username": f"api:{row['name']}",
        "role": "admin",
        "is_active": 1,
        "must_change_password": 0,
    }
    try:
        with transaction(conn):
            touch_api_key_last_used(conn, row["id"])
    except sqlite3.Error:
        current_app.logger.exception("failed to update api key last_used_at")
    return None


@api.after_request
def api_no_store(response: Response) -> Response:
    response.headers.setdefault("Cache-Control", "no-store")
    return response


@api.errorhandler(DomainError)
def handle_domain_error(error: DomainError):
    return _json_error(error.code, error.message, error.status)


@api.errorhandler(HTTPException)
def handle_http_error(error: HTTPException):
    code_map = {
        400: ("bad_request", "Requisição inválida"),
        401: ("unauthorized", "Autenticação necessária"),
        403: ("forbidden", "Acesso negado"),
        404: ("not_found", "Recurso não encontrado"),
        409: ("conflict", "Conflito"),
        413: ("payload_too_large", "Payload excede o limite"),
        429: ("rate_limited", "Muitas tentativas"),
        503: ("service_unavailable", "Serviço temporariamente indisponível"),
    }
    code, message = code_map.get(error.code or 500, ("error", "Erro"))
    response = _json_error(code, message, error.code or 500)
    if error.code == 401:
        response.headers["WWW-Authenticate"] = 'Bearer realm="service-manager"'
    return response


# ---- API keys ----


@api.get("/api-keys")
def api_keys_list() -> ResponseReturnValue:
    rows = list_api_keys(get_db())
    return _json(
        {
            "items": [
                {
                    "id": row["id"],
                    "name": row["name"],
                    "created_at": row["created_at"],
                    "last_used_at": row["last_used_at"],
                    "revoked_at": row["revoked_at"],
                    "active": row["revoked_at"] is None,
                }
                for row in rows
            ]
        }
    )


@api.post("/api-keys")
def api_keys_create() -> ResponseReturnValue:
    body = _json_body()
    conn = get_db()
    with transaction(conn):
        key_id, token = create_api_key(conn, _principal(), str(body.get("name") or ""))
    response = _json({"id": key_id, "api_key": token}, 201)
    response.headers["Cache-Control"] = "no-store, private"
    return response


@api.delete("/api-keys/<int:key_id>")
def api_keys_revoke(key_id: int) -> ResponseReturnValue:
    conn = get_db()
    with transaction(conn):
        revoke_api_key(conn, _principal(), key_id)
    return _empty()


# ---- Services / accounts ----


@api.get("/services")
def services_list() -> ResponseReturnValue:
    rows = get_db().execute("SELECT id, name, rotation_days FROM services ORDER BY name COLLATE NOCASE, id").fetchall()
    return _json({"items": [{"id": r["id"], "name": r["name"], "rotation_days": r["rotation_days"]} for r in rows]})


@api.post("/services")
def services_create() -> ResponseReturnValue:
    body = _json_body()
    conn = get_db()
    with transaction(conn):
        service_id = create_service(conn, _principal(), str(body.get("name") or ""))
    return _json({"id": service_id}, 201)


@api.get("/services/<int:service_id>")
def services_get(service_id: int) -> ResponseReturnValue:
    conn = get_db()
    require_service_role(conn, service_id, "viewer")
    row = conn.execute("SELECT id, name, rotation_days FROM services WHERE id=?", (service_id,)).fetchone()
    if row is None:
        raise NotFoundError()
    return _json({"id": row["id"], "name": row["name"], "rotation_days": row["rotation_days"]})


@api.delete("/services/<int:service_id>")
def services_delete(service_id: int) -> ResponseReturnValue:
    conn = get_db()
    with transaction(conn):
        delete_service(conn, _principal(), service_id)
    return _empty()


@api.patch("/services/<int:service_id>/rotation-policy")
def services_rotation_policy(service_id: int) -> ResponseReturnValue:
    body = _json_body()
    if "rotation_days" not in body:
        raise DomainError("invalid_rotation", "Intervalo inválido")
    conn = get_db()
    _require_rotation_enabled(conn)
    if conn.execute("SELECT 1 FROM services WHERE id=?", (service_id,)).fetchone() is None:
        raise NotFoundError()
    require_service_role(conn, service_id, "service_admin")
    days = _parse_rotation_days_value(body.get("rotation_days"))
    with transaction(conn):
        set_service_rotation_policy(conn, _principal(), service_id=service_id, rotation_days=days)
    return _empty()


@api.get("/services/<int:service_id>/accounts")
def accounts_list(service_id: int) -> ResponseReturnValue:
    conn = get_db()
    require_service_role(conn, service_id, "viewer")
    service = conn.execute("SELECT rotation_days FROM services WHERE id=?", (service_id,)).fetchone()
    if service is None:
        raise NotFoundError()
    q = (request.args.get("q") or "").strip()
    status = normalize_status(request.args.get("status"))
    registered_raw = request.args.get("registered")
    registered = None
    if registered_raw is not None:
        if registered_raw not in {"0", "1"}:
            raise DomainError("invalid_registered", "Cadastro inválido")
        registered = int(registered_raw)
    rotation = (request.args.get("rotation") or "").strip() or None
    sort = request.args.get("sort") or "email"
    if sort not in {"email", "status"}:
        sort = "email"
    direction = (request.args.get("direction") or "asc").lower()
    sql_dir = "DESC" if direction == "desc" else "ASC"
    op = "<" if sql_dir == "DESC" else ">"
    today = datetime.now(UTC).date()
    where, params = _account_filter_sql(
        service_id=service_id,
        q=q,
        status=status,
        registered=registered,
        rot_state=rotation,
        service_days=service["rotation_days"],
        today=today,
    )
    cursor_token = request.args.get("cursor") or ""
    if cursor_token:
        cursor = _decode_cursor("accounts", cursor_token)
        where = f"{where} AND {_account_keyset_sql(sort, op)}"
        params.update({"cursor_email": cursor.get("email"), "cursor_id": cursor.get("id"), "cursor_rank": cursor.get("rank")})
    order = _account_order_sql(sort, sql_dir)
    sql = f"{_ACCOUNT_SELECT} WHERE {where} ORDER BY {order} LIMIT {_ACCOUNT_PAGE_SIZE + 1}"
    rows = conn.execute(sql, params).fetchall()
    counts_row = conn.execute(
        f"""
        SELECT COUNT(*) AS total,
          SUM(CASE WHEN link.status='ativo' THEN 1 ELSE 0 END) AS ativo,
          SUM(CASE WHEN link.status='nunca' THEN 1 ELSE 0 END) AS nunca,
          SUM(CASE WHEN link.status='inativo' THEN 1 ELSE 0 END) AS inativo,
          SUM(CASE WHEN link.registered=1 THEN 1 ELSE 0 END) AS registered
        FROM account_service AS link
        JOIN accounts AS a ON a.id = link.account_id
        WHERE {where}
        """,
        params,
    ).fetchone()
    items = []
    for row in rows[:_ACCOUNT_PAGE_SIZE]:
        state = _rotation_state(row["password_changed_at"], row["rotation_days"], row["rotation_due_at"], service["rotation_days"], today=today)
        items.append(
            {
                "id": row["id"],
                "email": row["email"],
                "status": row["status"],
                "registered": bool(row["registered"]),
                "password_changed_at": row["password_changed_at"],
                "rotation_days": row["rotation_days"],
                "rotation_due_at": row["rotation_due_at"],
                "rotation": state,
            }
        )
    next_cursor = None
    if len(rows) > _ACCOUNT_PAGE_SIZE:
        last = rows[_ACCOUNT_PAGE_SIZE - 1]
        payload: dict[str, object] = {"email": last["email"], "id": last["id"]}
        if sort == "status":
            payload["rank"] = {"ativo": 0, "nunca": 1, "inativo": 2}.get(last["status"], 1)
        next_cursor = _encode_cursor("accounts", payload)
    return _json(
        {
            "items": items,
            "counts": {
                "total": counts_row["total"] or 0,
                "ativo": counts_row["ativo"] or 0,
                "nunca": counts_row["nunca"] or 0,
                "inativo": counts_row["inativo"] or 0,
                "registered": counts_row["registered"] or 0,
            },
            "next_cursor": next_cursor,
            "previous_cursor": None,
        }
    )


@api.post("/services/<int:service_id>/accounts")
def accounts_create(service_id: int) -> ResponseReturnValue:
    body = _json_body()
    conn = get_db()
    require_service_role(conn, service_id, "editor")
    registered = body.get("registered", False)
    if not isinstance(registered, bool):
        raise DomainError("invalid_registered", "Cadastro inválido")
    with transaction(conn):
        account_id = create_account(
            conn,
            _principal(),
            service_id=service_id,
            email=str(body.get("email") or ""),
            password=str(body.get("password") or ""),
            status=str(body.get("status") or ""),
            registered=1 if registered else 0,
        )
    return _json({"id": account_id}, 201)


@api.get("/services/<int:service_id>/accounts/<int:account_id>")
def accounts_get(service_id: int, account_id: int) -> ResponseReturnValue:
    conn = get_db()
    require_account_role(conn, account_id, service_id, "viewer")
    row = conn.execute(
        """
        SELECT a.id, a.email, a.password_changed_at, link.status, link.registered, link.rotation_days, link.rotation_due_at
        FROM accounts AS a JOIN account_service AS link ON link.account_id = a.id
        WHERE a.id=? AND link.service_id=?
        """,
        (account_id, service_id),
    ).fetchone()
    if row is None:
        raise NotFoundError()
    service = conn.execute("SELECT rotation_days FROM services WHERE id=?", (service_id,)).fetchone()
    fields = []
    for field in conn.execute(
        """
        SELECT f.id, f.name, v.value_ciphertext, v.value_nonce, v.value_key_version
        FROM custom_fields AS f
        LEFT JOIN field_values AS v ON v.field_id = f.id AND v.account_id = ?
        WHERE f.service_id = ?
        ORDER BY f.name COLLATE NOCASE, f.id
        """,
        (account_id, service_id),
    ):
        value = None
        if field["value_ciphertext"] is not None:
            value = decrypt_secret(
                EncryptedValue(field["value_ciphertext"], field["value_nonce"], field["value_key_version"]),
                aad=account_field_aad(account_id, field["id"]),
            )
        fields.append({"id": field["id"], "name": field["name"], "value": value})
    state = _rotation_state(
        row["password_changed_at"],
        row["rotation_days"],
        row["rotation_due_at"],
        service["rotation_days"] if service else None,
        today=datetime.now(UTC).date(),
    )
    return _json(
        {
            "id": row["id"],
            "email": row["email"],
            "status": row["status"],
            "registered": bool(row["registered"]),
            "password_changed_at": row["password_changed_at"],
            "rotation_days": row["rotation_days"],
            "rotation_due_at": row["rotation_due_at"],
            "rotation": state,
            "fields": fields,
        }
    )


@api.patch("/services/<int:service_id>/accounts/<int:account_id>")
def accounts_update(service_id: int, account_id: int) -> ResponseReturnValue:
    body = _json_body()
    conn = get_db()
    require_account_role(conn, account_id, service_id, "editor")
    with transaction(conn):
        update_account(
            conn,
            _principal(),
            account_id=account_id,
            service_id=service_id,
            email=str(body.get("email") or ""),
            password=body.get("password") if "password" in body else "",
        )
    return _empty()


@api.delete("/services/<int:service_id>/accounts/<int:account_id>")
def accounts_delete(service_id: int, account_id: int) -> ResponseReturnValue:
    conn = get_db()
    require_account_role(conn, account_id, service_id, "service_admin", all_linked_services=True)
    with transaction(conn):
        delete_account(conn, _principal(), account_id=account_id, service_id=service_id)
    return _empty()


@api.post("/services/<int:service_id>/accounts/<int:account_id>/password/reveal")
def accounts_reveal_password(service_id: int, account_id: int) -> ResponseReturnValue:
    conn = get_db()
    require_account_role(conn, account_id, service_id, "editor")
    with transaction(conn):
        value = reveal_account_password(
            conn,
            _principal(),
            account_id=account_id,
            subject=f"api-key:{g.api_key['id']}",
            source_ip=source_ip(),
        )
    response = _json({"value": value, "expires_in": 30})
    response.headers["Cache-Control"] = "no-store, private"
    return response


@api.patch("/services/<int:service_id>/accounts/<int:account_id>/status")
def accounts_status(service_id: int, account_id: int) -> ResponseReturnValue:
    body = _json_body()
    conn = get_db()
    require_account_role(conn, account_id, service_id, "editor")
    with transaction(conn):
        update_account_status(conn, _principal(), account_id=account_id, service_id=service_id, status=str(body.get("status") or ""))
    return _empty()


@api.patch("/services/<int:service_id>/accounts/<int:account_id>/registered")
def accounts_registered(service_id: int, account_id: int) -> ResponseReturnValue:
    body = _json_body()
    registered = body.get("registered")
    if not isinstance(registered, bool):
        raise DomainError("invalid_registered", "Cadastro inválido")
    conn = get_db()
    require_account_role(conn, account_id, service_id, "editor")
    with transaction(conn):
        update_account_registered(conn, _principal(), account_id=account_id, service_id=service_id, registered=1 if registered else 0)
    return _empty()


@api.patch("/services/<int:service_id>/accounts/<int:account_id>/rotation-policy")
def accounts_rotation_policy(service_id: int, account_id: int) -> ResponseReturnValue:
    body = _json_body()
    conn = get_db()
    _require_rotation_enabled(conn)
    require_account_role(conn, account_id, service_id, "editor")
    if "rotation_days" not in body and "rotation_due_at" not in body:
        raise DomainError("invalid_rotation", "Política de rotação inválida")
    days = _parse_rotation_days_value(body["rotation_days"]) if "rotation_days" in body else None
    due_at = _parse_rotation_due_at_value(body["rotation_due_at"]) if "rotation_due_at" in body else None
    with transaction(conn):
        set_account_rotation_policy(
            conn,
            _principal(),
            account_id=account_id,
            service_id=service_id,
            rotation_days=days if "rotation_days" in body else None,
            rotation_due_at=due_at if "rotation_due_at" in body else None,
        )
    return _empty()


@api.post("/services/<int:service_id>/accounts/<int:account_id>/rotation")
def accounts_rotation(service_id: int, account_id: int) -> ResponseReturnValue:
    body = _json_body()
    conn = get_db()
    _require_rotation_enabled(conn)
    require_account_role(conn, account_id, service_id, "editor")
    with transaction(conn):
        complete_rotation(
            conn,
            _principal(),
            account_id=account_id,
            service_id=service_id,
            outcome=str(body.get("outcome") or ""),
            new_password=body.get("new_password") if isinstance(body.get("new_password"), str) else None,
        )
    return _empty()


@api.get("/services/<int:service_id>/accounts/<int:account_id>/fields/<int:field_id>")
def fields_get(service_id: int, account_id: int, field_id: int) -> ResponseReturnValue:
    conn = get_db()
    require_account_role(conn, account_id, service_id, "viewer")
    field = conn.execute("SELECT id, name FROM custom_fields WHERE id=? AND service_id=?", (field_id, service_id)).fetchone()
    if field is None:
        raise NotFoundError()
    value_row = conn.execute(
        "SELECT value_ciphertext, value_nonce, value_key_version FROM field_values WHERE field_id=? AND account_id=?",
        (field_id, account_id),
    ).fetchone()
    value = None
    if value_row is not None:
        value = decrypt_secret(
            EncryptedValue(value_row["value_ciphertext"], value_row["value_nonce"], value_row["value_key_version"]),
            aad=account_field_aad(account_id, field_id),
        )
    return _json({"id": field["id"], "name": field["name"], "value": value})


@api.put("/services/<int:service_id>/accounts/<int:account_id>/fields/<int:field_id>")
def fields_put(service_id: int, account_id: int, field_id: int) -> ResponseReturnValue:
    body = _json_body()
    conn = get_db()
    require_account_role(conn, account_id, service_id, "editor")
    with transaction(conn):
        update_field_value(
            conn,
            _principal(),
            service_id=service_id,
            field_id=field_id,
            account_id=account_id,
            value=str(body.get("value") if body.get("value") is not None else ""),
        )
    return _empty()


@api.delete("/services/<int:service_id>/accounts/<int:account_id>/fields/<int:field_id>")
def fields_delete(service_id: int, account_id: int, field_id: int) -> ResponseReturnValue:
    conn = get_db()
    require_account_role(conn, account_id, service_id, "service_admin")
    with transaction(conn):
        delete_field_value(conn, _principal(), service_id=service_id, field_id=field_id, account_id=account_id)
    return _empty()


@api.post("/services/<int:service_id>/fields")
def fields_create(service_id: int) -> ResponseReturnValue:
    body = _json_body()
    account_ids = _bulk_ids(body.get("account_ids"))
    conn = get_db()
    for account_id in account_ids:
        require_account_role(conn, account_id, service_id, "editor")
    with transaction(conn):
        field_id = upsert_field_values(
            conn,
            _principal(),
            service_id=service_id,
            name=str(body.get("name") or ""),
            account_ids=account_ids,
            value=str(body.get("value") if body.get("value") is not None else ""),
        )
    return _json({"id": field_id}, 201)


@api.post("/services/<int:service_id>/accounts/bulk/status")
def bulk_status(service_id: int) -> ResponseReturnValue:
    body = _json_body()
    account_ids = _bulk_ids(body.get("account_ids"))
    conn = get_db()
    require_accounts_role(conn, account_ids, service_id, "editor")
    with transaction(conn):
        bulk_update_status(conn, _principal(), service_id=service_id, account_ids=account_ids, status=str(body.get("status") or ""))
    return _empty()


@api.post("/services/<int:service_id>/accounts/bulk/registered")
def bulk_registered(service_id: int) -> ResponseReturnValue:
    body = _json_body()
    account_ids = _bulk_ids(body.get("account_ids"))
    registered = body.get("registered")
    if not isinstance(registered, bool):
        raise DomainError("invalid_registered", "Cadastro inválido")
    conn = get_db()
    require_accounts_role(conn, account_ids, service_id, "editor")
    with transaction(conn):
        bulk_update_registered(conn, _principal(), service_id=service_id, account_ids=account_ids, registered=1 if registered else 0)
    return _empty()


@api.post("/services/<int:service_id>/accounts/bulk/field")
def bulk_field(service_id: int) -> ResponseReturnValue:
    body = _json_body()
    account_ids = _bulk_ids(body.get("account_ids"))
    try:
        raw_field_id = body.get("field_id")
        if raw_field_id is None or isinstance(raw_field_id, bool):
            raise DomainError("invalid_field", "Campo inválido")
        field_id = int(raw_field_id)
    except (TypeError, ValueError) as error:
        raise DomainError("invalid_field", "Campo inválido") from error
    conn = get_db()
    require_accounts_role(conn, account_ids, service_id, "editor")
    with transaction(conn):
        bulk_set_field(
            conn,
            _principal(),
            service_id=service_id,
            account_ids=account_ids,
            field_id=field_id,
            value=str(body.get("value") if body.get("value") is not None else ""),
        )
    return _empty()


@api.post("/services/<int:service_id>/accounts/bulk/fields")
def bulk_fields(service_id: int) -> ResponseReturnValue:
    body = _json_body()
    account_ids = _bulk_ids(body.get("account_ids"))
    conn = get_db()
    require_accounts_role(conn, account_ids, service_id, "editor")
    with transaction(conn):
        field_id, created_count = bulk_create_field(
            conn,
            _principal(),
            service_id=service_id,
            account_ids=account_ids,
            name=str(body.get("name") or body.get("field_name") or ""),
        )
    return _json({"field_id": field_id, "created_count": created_count})


@api.post("/services/<int:service_id>/accounts/bulk/delete")
def bulk_delete(service_id: int) -> ResponseReturnValue:
    body = _json_body()
    account_ids = _bulk_ids(body.get("account_ids"))
    conn = get_db()
    require_accounts_role(conn, account_ids, service_id, "service_admin", all_linked_services=True)
    with transaction(conn):
        bulk_delete_accounts(conn, _principal(), service_id=service_id, account_ids=account_ids)
    return _empty()


# ---- import/export/rotation/coverage ----


@api.get("/services/<int:service_id>/imports/template.csv")
def import_template_csv(service_id: int) -> ResponseReturnValue:
    conn = get_db()
    require_service_role(conn, service_id, "viewer")
    stream = io.StringIO()
    csv.writer(stream).writerows(TEMPLATE_ROWS)
    return _no_store(Response(stream.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=modelo_credenciais.csv"}))


@api.get("/services/<int:service_id>/imports/template.xlsx")
def import_template_xlsx(service_id: int) -> ResponseReturnValue:
    from openpyxl import Workbook

    conn = get_db()
    require_service_role(conn, service_id, "viewer")
    workbook = Workbook()
    worksheet = workbook.active
    if worksheet is None:
        raise RuntimeError("new workbook has no active worksheet")
    for row in TEMPLATE_ROWS:
        worksheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    return _no_store(
        Response(
            stream.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=modelo_credenciais.xlsx"},
        )
    )


@api.post("/services/<int:service_id>/imports")
def import_accounts_endpoint(service_id: int) -> ResponseReturnValue:
    conn = get_db()
    require_service_role(conn, service_id, "service_admin")
    upload = request.files.get("file")
    if upload is None or not upload.filename:
        raise DomainError("invalid_import", "Arquivo inválido")
    if not has_allowed_upload_mimetype(upload.filename, upload.mimetype):
        raise DomainError("invalid_import", "Formato inválido")
    try:
        records = parse_import_file(upload.filename, upload.stream)
    except ImportFormatError as error:
        raise DomainError("invalid_import", "Formato inválido") from error
    normalized_records: list[tuple[str, str, str, tuple[str, ...]]] = []
    for record in records.records:
        normalized_email = valid_email(record.email)
        normalized_status = normalize_status(record.status)
        if normalized_email is None or normalized_status is None or valid_secret(record.password) is None:
            raise DomainError("invalid_import", "Validação falhou")
        if any(valid_secret(value) is None for value in record.field_values):
            raise DomainError("invalid_import", "Validação falhou")
        normalized_records.append((normalized_email, record.password, normalized_status, record.field_values))
    if any(valid_name(name) is None for name in records.field_names):
        raise DomainError("invalid_import", "Validação falhou")
    with transaction(conn):
        added, skipped = import_accounts(
            conn,
            _principal(),
            service_id=service_id,
            records=normalized_records,
            field_names=records.field_names,
        )
    return _json({"added": added, "skipped": skipped})


def _export_common(service_id: int):
    conn = get_db()
    require_service_role(conn, service_id, "service_admin")
    count = conn.execute("SELECT COUNT(*) AS n FROM account_service WHERE service_id = ?", (service_id,)).fetchone()["n"]
    if count > _EXPORT_LIMIT:
        raise DomainError("export_limit", "Exportação limitada a 10000 contas.", 413)
    service_name = conn.execute("SELECT name FROM services WHERE id=?", (service_id,)).fetchone()
    if service_name is None:
        raise NotFoundError()
    return conn, count, service_name["name"]


@api.get("/services/<int:service_id>/exports/accounts.csv")
def export_accounts_csv(service_id: int) -> ResponseReturnValue:
    conn, count, name = _export_common(service_id)
    stamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    filename = f"contas_{_safe_filename_slug(name)}_{service_id}_{stamp}.csv"
    fields = _export_fields(conn, service_id)
    headers = ("email", "password", "status", "cadastrada", *(_sanitize_cell(f"campo:{n}") for _, n in fields))
    spool = tempfile.SpooledTemporaryFile(max_size=8 * 1024 * 1024)
    try:
        buffer = io.StringIO(newline="")
        writer = csv.writer(buffer)

        def _flush() -> None:
            spool.write(buffer.getvalue().encode("utf-8"))
            buffer.seek(0)
            buffer.truncate(0)

        spool.write("\ufeff".encode("utf-8"))
        writer.writerow(headers)
        _flush()
        for row in _iter_export_rows(conn, service_id, fields):
            writer.writerow(row)
            _flush()
        spool.seek(0)
        response = send_file(spool, mimetype="text/csv", as_attachment=True, download_name=filename)
        response.call_on_close(spool.close)
        with transaction(conn):
            domain_audit(conn, _principal(), action="accounts.exported", target_type="service", target_id=service_id, metadata={"rows": count, "format": "csv"})
        return _no_store(response)
    except Exception:
        spool.close()
        raise


@api.get("/services/<int:service_id>/exports/accounts.xlsx")
def export_accounts_xlsx(service_id: int) -> ResponseReturnValue:
    from openpyxl import Workbook

    conn, count, name = _export_common(service_id)
    stamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    filename = f"contas_{_safe_filename_slug(name)}_{service_id}_{stamp}.xlsx"
    fields = _export_fields(conn, service_id)
    headers = ("email", "password", "status", "cadastrada", *(_sanitize_cell(f"campo:{n}") for _, n in fields))
    spool = tempfile.SpooledTemporaryFile(max_size=8 * 1024 * 1024)
    try:
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet()
        worksheet.append(headers)
        for row in _iter_export_rows(conn, service_id, fields):
            worksheet.append(row)
        workbook.save(spool)
        spool.seek(0)
        response = send_file(
            spool,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
        response.call_on_close(spool.close)
        with transaction(conn):
            domain_audit(conn, _principal(), action="accounts.exported", target_type="service", target_id=service_id, metadata={"rows": count, "format": "xlsx"})
        return _no_store(response)
    except Exception:
        spool.close()
        raise


@api.get("/services/<int:service_id>/rotation")
def rotation_list(service_id: int) -> ResponseReturnValue:
    conn = get_db()
    _require_rotation_enabled(conn)
    require_service_role(conn, service_id, "viewer")
    service_row = conn.execute("SELECT name, rotation_days FROM services WHERE id=?", (service_id,)).fetchone()
    if service_row is None:
        raise NotFoundError()
    today = datetime.now(UTC).date()
    accounts = []
    for row in conn.execute(
        """
        SELECT a.id, a.email, a.password_changed_at, link.rotation_days, link.rotation_due_at
        FROM account_service AS link JOIN accounts AS a ON a.id = link.account_id
        WHERE link.service_id = ? ORDER BY a.email COLLATE NOCASE
        """,
        (service_id,),
    ):
        state = _rotation_state(row["password_changed_at"], row["rotation_days"], row["rotation_due_at"], service_row["rotation_days"], today=today)
        if state["state"] in ("unknown", "due_soon", "overdue"):
            accounts.append(
                {
                    "id": row["id"],
                    "email": row["email"],
                    "password_changed_at": row["password_changed_at"],
                    "effective_days": state["effective_days"],
                    "due_at": state["due_at"],
                    "days_remaining": state["days_remaining"],
                    "state": state["state"],
                }
            )
    return _json({"service_id": service_id, "service_name": service_row["name"], "items": accounts})


@api.get("/coverage")
def coverage() -> ResponseReturnValue:
    conn = get_db()
    services = accessible_services(conn, g.current_user)
    service_ids = [row["id"] for row in services]
    if not service_ids:
        return _json({"items": [], "next_cursor": None, "previous_cursor": None, "services": []})
    filt = (request.args.get("filter") or "").strip()
    if filt not in {"", "none-registered", "multi-active", "missing-registration"}:
        raise DomainError("invalid_filter", "Filtro inválido")
    placeholders = ",".join("?" for _ in service_ids)
    accounts = conn.execute(
        f"""
        SELECT a.id, a.email FROM accounts AS a
        WHERE EXISTS (SELECT 1 FROM account_service AS link WHERE link.account_id = a.id AND link.service_id IN ({placeholders}))
        ORDER BY a.email COLLATE NOCASE, a.id LIMIT ?
        """,
        (*service_ids, _COVERAGE_PAGE_SIZE + 1),
    ).fetchall()
    links = conn.execute(
        f"SELECT account_id, service_id, status, registered FROM account_service WHERE service_id IN ({placeholders})",
        service_ids,
    ).fetchall()
    by_account: dict[int, dict[int, dict[str, Any]]] = {}
    for link in links:
        by_account.setdefault(link["account_id"], {})[link["service_id"]] = {
            "status": link["status"],
            "registered": bool(link["registered"]),
        }
    items = []
    for account in accounts[:_COVERAGE_PAGE_SIZE]:
        matrix = by_account.get(account["id"], {})
        if filt == "none-registered" and any(cell.get("registered") for cell in matrix.values()):
            continue
        if filt == "multi-active" and sum(1 for cell in matrix.values() if cell.get("status") == "ativo") < 2:
            continue
        if filt == "missing-registration" and not any(cell.get("status") == "ativo" and not cell.get("registered") for cell in matrix.values()):
            continue
        items.append(
            {
                "id": account["id"],
                "email": account["email"],
                "services": {
                    str(service_id): matrix.get(service_id, {"status": "nunca", "registered": False})
                    for service_id in service_ids
                },
            }
        )
    return _json(
        {
            "items": items,
            "services": [{"id": row["id"], "name": row["name"]} for row in services],
            "next_cursor": None,
            "previous_cursor": None,
        }
    )


# ---- users/memberships/preferences ----


@api.get("/users")
def users_list() -> ResponseReturnValue:
    rows = get_db().execute(
        "SELECT id, username, role, is_active, must_change_password, created_at, updated_at FROM users ORDER BY username"
    ).fetchall()
    return _json(
        {
            "items": [
                {
                    "id": row["id"],
                    "username": row["username"],
                    "role": row["role"],
                    "is_active": bool(row["is_active"]),
                    "must_change_password": bool(row["must_change_password"]),
                    "created_at": row["created_at"],
                    "updated_at": row["updated_at"],
                }
                for row in rows
            ]
        }
    )


@api.post("/users")
def users_create() -> ResponseReturnValue:
    body = _json_body()
    conn = get_db()
    with transaction(conn):
        user_id, temporary_password = create_user(conn, _principal(), username=str(body.get("username") or ""), role=str(body.get("role") or ""))
    response = _json({"id": user_id, "temporary_password": temporary_password}, 201)
    response.headers["Cache-Control"] = "no-store, private"
    return response


@api.get("/users/<int:user_id>")
def users_get(user_id: int) -> ResponseReturnValue:
    row = get_db().execute(
        "SELECT id, username, role, is_active, must_change_password, created_at, updated_at FROM users WHERE id=?",
        (user_id,),
    ).fetchone()
    if row is None:
        raise NotFoundError()
    return _json(
        {
            "id": row["id"],
            "username": row["username"],
            "role": row["role"],
            "is_active": bool(row["is_active"]),
            "must_change_password": bool(row["must_change_password"]),
            "created_at": row["created_at"],
            "updated_at": row["updated_at"],
        }
    )


@api.patch("/users/<int:user_id>")
def users_patch(user_id: int) -> ResponseReturnValue:
    body = _json_body()
    if not any(key in body for key in ("role", "is_active")) or set(body) - {"role", "is_active"}:
        raise DomainError("invalid_user", "Usuário inválido")
    conn = get_db()
    with transaction(conn):
        if "role" in body:
            change_user_role(conn, _principal(), user_id=user_id, role=str(body.get("role") or ""))
        if "is_active" in body:
            if not isinstance(body.get("is_active"), bool):
                raise DomainError("invalid_user", "Usuário inválido")
            change_user_active(conn, _principal(), user_id=user_id, is_active=bool(body["is_active"]))
    return _empty()


@api.get("/memberships")
def memberships_list() -> ResponseReturnValue:
    conn = get_db()
    clauses: list[str] = []
    params: list[object] = []
    if request.args.get("user_id"):
        try:
            params.append(int(request.args["user_id"]))
        except ValueError as error:
            raise DomainError("invalid_filter", "Filtro inválido") from error
        clauses.append("user_id = ?")
    if request.args.get("service_id"):
        try:
            params.append(int(request.args["service_id"]))
        except ValueError as error:
            raise DomainError("invalid_filter", "Filtro inválido") from error
        clauses.append("service_id = ?")
    where = f" WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = conn.execute(
        f"SELECT user_id, service_id, role, created_at FROM service_members{where} ORDER BY service_id, user_id",
        params,
    ).fetchall()
    return _json(
        {
            "items": [
                {"user_id": row["user_id"], "service_id": row["service_id"], "role": row["role"], "created_at": row["created_at"]}
                for row in rows
            ]
        }
    )


@api.put("/services/<int:service_id>/members/<int:user_id>")
def memberships_put(service_id: int, user_id: int) -> ResponseReturnValue:
    body = _json_body()
    conn = get_db()
    with transaction(conn):
        grant_membership(conn, _principal(), service_id=service_id, user_id=user_id, role=str(body.get("role") or ""))
    return _empty()


@api.delete("/services/<int:service_id>/members/<int:user_id>")
def memberships_delete(service_id: int, user_id: int) -> ResponseReturnValue:
    conn = get_db()
    with transaction(conn):
        revoke_membership(conn, _principal(), service_id=service_id, user_id=user_id)
    return _empty()


@api.get("/users/<int:user_id>/service-preferences")
def preferences_get(user_id: int) -> ResponseReturnValue:
    conn = get_db()
    if conn.execute("SELECT id FROM users WHERE id=?", (user_id,)).fetchone() is None:
        raise NotFoundError()
    rows = conn.execute(
        "SELECT service_id, position, is_initial FROM user_service_preferences WHERE user_id=? ORDER BY position",
        (user_id,),
    ).fetchall()
    return _json(
        {
            "service_ids": [row["service_id"] for row in rows],
            "initial_service_id": next((row["service_id"] for row in rows if row["is_initial"]), None),
        }
    )


@api.put("/users/<int:user_id>/service-preferences")
def preferences_put(user_id: int) -> ResponseReturnValue:
    body = _json_body()
    conn = get_db()
    user = conn.execute("SELECT id, role, is_active FROM users WHERE id=?", (user_id,)).fetchone()
    if user is None:
        raise NotFoundError()
    raw_ids = body.get("service_ids")
    if not isinstance(raw_ids, list):
        raise DomainError("invalid_preferences", "Preferências de serviços inválidas")
    try:
        service_ids = [int(value) for value in raw_ids]
    except (TypeError, ValueError) as error:
        raise DomainError("invalid_preferences", "Preferências de serviços inválidas") from error
    if len(service_ids) != len(set(service_ids)):
        raise DomainError("invalid_preferences", "Preferências de serviços inválidas")
    initial = body.get("initial_service_id")
    if initial is not None and not isinstance(initial, int):
        raise DomainError("invalid_preferences", "Preferências de serviços inválidas")
    services = accessible_services(conn, user)
    accessible_ids = [service["id"] for service in services]
    if sorted(service_ids) != sorted(accessible_ids):
        raise DomainError("invalid_preferences", "Preferências de serviços inválidas")
    if initial is not None and initial not in service_ids:
        raise DomainError("invalid_preferences", "Preferências de serviços inválidas")
    with transaction(conn):
        update_service_preferences(conn, _principal(), user_id=user_id, service_ids=service_ids, initial_service_id=initial)
    return _empty()


# ---- webhooks/settings/audit ----


@api.get("/webhooks")
def webhooks_list() -> ResponseReturnValue:
    configs = list_webhook_configs(get_db())
    return _json({"items": configs, "event_types": list(webhook_event_types()), "at_capacity": len(configs) >= 20})


@api.post("/webhooks")
def webhooks_create() -> ResponseReturnValue:
    body = _json_body()
    url = str(body.get("url") or "").strip()
    description = str(body.get("description") or "").strip()
    enabled = body.get("enabled", True)
    event_types = body.get("event_types") or []
    if not isinstance(enabled, bool) or not isinstance(event_types, list):
        raise DomainError("invalid_webhook", "Integração inválida")
    conn = get_db()
    with transaction(conn):
        config_id, host, secret, subscriptions = create_webhook(
            conn,
            _principal(),
            url=url,
            description=description,
            enabled=enabled,
            event_types=[str(item) for item in event_types],
            data_key_b64=current_app.config["DATA_KEY_V1"],
            resolver=_webhook_resolver(),
        )
    response = _json({"id": config_id, "signing_secret": secret}, 201)
    response.headers["Cache-Control"] = "no-store, private"
    return response


@api.get("/webhooks/<int:config_id>")
def webhooks_get(config_id: int) -> ResponseReturnValue:
    for config in list_webhook_configs(get_db()):
        if config.get("id") == config_id:
            return _json(config)
    raise NotFoundError()


@api.patch("/webhooks/<int:config_id>")
def webhooks_patch(config_id: int) -> ResponseReturnValue:
    body = _json_body()
    url = str(body.get("url") or "").strip()
    description = str(body.get("description") or "").strip()
    enabled = body.get("enabled", True)
    event_types = body.get("event_types") or []
    if not isinstance(enabled, bool) or not isinstance(event_types, list):
        raise DomainError("invalid_webhook", "Integração inválida")
    conn = get_db()
    with transaction(conn):
        update_webhook(
            conn,
            _principal(),
            config_id=config_id,
            url=url,
            description=description,
            enabled=enabled,
            event_types=[str(item) for item in event_types],
            data_key_b64=current_app.config["DATA_KEY_V1"],
            resolver=_webhook_resolver(),
        )
    return _empty()


@api.delete("/webhooks/<int:config_id>")
def webhooks_delete(config_id: int) -> ResponseReturnValue:
    conn = get_db()
    with transaction(conn):
        delete_webhook(conn, _principal(), config_id=config_id)
    return _empty()


@api.post("/webhooks/<int:config_id>/test")
def webhooks_test(config_id: int) -> ResponseReturnValue:
    conn = get_db()
    with transaction(conn):
        test_webhook(conn, _principal(), config_id=config_id)
    return _empty()


@api.get("/settings")
def settings_get() -> ResponseReturnValue:
    row = get_db().execute("SELECT value FROM app_settings WHERE key='rotation_enabled'").fetchone()
    return _json({"rotation_enabled": row is not None and row["value"] == "1"})


@api.patch("/settings")
def settings_patch() -> ResponseReturnValue:
    body = _json_body()
    if "rotation_enabled" not in body or not isinstance(body.get("rotation_enabled"), bool):
        raise DomainError("invalid_settings", "Configuração inválida")
    conn = get_db()
    with transaction(conn):
        set_rotation_enabled_setting(conn, _principal(), enabled=bool(body["rotation_enabled"]))
    return _empty()


def _audit_query_filters() -> tuple[str, list[object], dict[str, str]]:
    filters = {name: (request.args.get(name) or "").strip() for name in ("action", "target_type", "actor", "api_key", "since", "until", "source_ip")}
    clauses: list[str] = []
    params: list[object] = []
    if filters["action"]:
        clauses.append("e.action = ?")
        params.append(filters["action"])
    if filters["target_type"]:
        clauses.append("e.target_type = ?")
        params.append(filters["target_type"])
    if filters["actor"]:
        try:
            actor = int(filters["actor"])
        except ValueError:
            actor = 0
        if actor > 0:
            clauses.append("e.actor_user_id = ?")
            params.append(actor)
        else:
            filters["actor"] = ""
    if filters["api_key"]:
        try:
            api_key_id = int(filters["api_key"])
        except ValueError:
            api_key_id = 0
        if api_key_id > 0:
            clauses.append("e.metadata_json LIKE ?")
            params.append(f'%"api_key_id": {api_key_id}%')
        else:
            filters["api_key"] = ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", filters["since"]):
        try:
            since = date.fromisoformat(filters["since"])
        except ValueError:
            filters["since"] = ""
        else:
            clauses.append("e.occurred_at >= ?")
            params.append(since.isoformat())
    else:
        filters["since"] = ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", filters["until"]):
        try:
            until = date.fromisoformat(filters["until"])
        except ValueError:
            filters["until"] = ""
        else:
            clauses.append("e.occurred_at < ?")
            params.append((until + timedelta(days=1)).isoformat())
    else:
        filters["until"] = ""
    if filters["source_ip"]:
        escaped = filters["source_ip"].replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        clauses.append("e.source_ip LIKE ? ESCAPE '\\\\'")
        params.append(f"%{escaped}%")
    return (" WHERE " + " AND ".join(clauses) if clauses else "", params, filters)


@api.get("/audit-events")
def audit_events() -> ResponseReturnValue:
    conn = get_db()
    try:
        page = max(1, int(request.args.get("page", "1")))
    except ValueError:
        page = 1
    where, params, filters = _audit_query_filters()
    rows = conn.execute(
        f"""
        SELECT e.id, e.occurred_at, e.action, e.target_type, e.target_id, e.metadata_json, e.source_ip, u.username
        FROM audit_events AS e LEFT JOIN users AS u ON u.id = e.actor_user_id
        {where}
        ORDER BY e.id DESC LIMIT 51 OFFSET ?
        """,
        (*params, (page - 1) * 50),
    ).fetchall()
    items = [
        {
            "id": row["id"],
            "occurred_at": row["occurred_at"],
            "action": row["action"],
            "target_type": row["target_type"],
            "target_id": row["target_id"],
            "metadata_json": row["metadata_json"],
            "source_ip": row["source_ip"],
            "actor_label": _actor_label(row["username"], row["metadata_json"]),
        }
        for row in rows[:50]
    ]
    return _json({"items": items, "page": page, "has_next": len(rows) > 50, "filters": filters, "chain_healthy": verify_audit_chain(conn)})


@api.get("/audit-events.csv")
def audit_events_csv() -> ResponseReturnValue:
    conn = get_db()
    where, params, _ = _audit_query_filters()
    cursor = conn.execute(
        f"""
        SELECT e.id, e.occurred_at, u.username, e.action, e.target_type, e.target_id, e.metadata_json, e.source_ip, e.previous_hash, e.event_hash
        FROM audit_events AS e LEFT JOIN users AS u ON u.id = e.actor_user_id
        {where}
        ORDER BY e.id DESC LIMIT 10000
        """,
        params,
    )
    filename = f"auditoria_{datetime.now(UTC).strftime('%Y%m%dT%H%M%SZ')}.csv"
    spool = tempfile.SpooledTemporaryFile(max_size=8 * 1024 * 1024)
    try:
        buffer = io.StringIO(newline="")
        writer = csv.writer(buffer)

        def _flush() -> None:
            spool.write(buffer.getvalue().encode("utf-8"))
            buffer.seek(0)
            buffer.truncate(0)

        spool.write("\ufeff".encode("utf-8"))
        writer.writerow(("id", "occurred_at", "usuario", "action", "target_type", "target_id", "metadata_json", "source_ip", "previous_hash", "event_hash"))
        _flush()
        for row in cursor:
            values = list(row)
            values[2] = _actor_label(row["username"], row["metadata_json"])
            values[8] = row["previous_hash"].hex() if row["previous_hash"] is not None else ""
            values[9] = row["event_hash"].hex() if row["event_hash"] is not None else ""
            writer.writerow(tuple(_sanitize_cell("" if value is None else str(value)) for value in values))
            _flush()
        spool.seek(0)
        response = send_file(spool, mimetype="text/csv", as_attachment=True, download_name=filename)
        response.call_on_close(spool.close)
        return _no_store(response)
    except Exception:
        spool.close()
        raise
