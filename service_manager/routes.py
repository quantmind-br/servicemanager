from __future__ import annotations

import csv
import io
import tempfile
import unicodedata
from datetime import UTC, datetime
from datetime import date, timedelta
import re
import sqlite3
import socket

from collections.abc import Iterator, Mapping
from typing import Any
from itsdangerous import BadSignature, URLSafeSerializer

from service_manager.auth import normalize_email, now_text, source_ip
from flask import Blueprint, Response, abort, current_app, g, jsonify, redirect, render_template, request, send_file, url_for
from flask.typing import ResponseReturnValue
from service_manager.audit import append_audit_event, verify_audit_chain

from service_manager.crypto import EncryptedValue, account_field_aad, account_password_aad, decrypt_secret, encrypt_secret
from service_manager.db import get_db, inserted_id, schema_is_current, transaction
from service_manager.imports import ImportFormatError, has_allowed_upload_mimetype, parse_import_file
from service_manager.authorization import (
    accessible_services,
    get_user_service_role,
    replace_service_preferences,
    require_account_role,
    require_accounts_role,
    require_recent_reauth,
    require_role,
    require_service_role,
    SERVICE_ROLE_RANK,
)
from service_manager.operations import (
    DomainError,
    NotFoundError,
    bulk_create_field,
    bulk_delete_accounts,
    bulk_set_field,
    bulk_update_registered,
    bulk_update_status,
    complete_rotation as op_complete_rotation,
    create_account,
    create_api_key,
    create_service,
    create_webhook,
    delete_account,
    delete_field_value,
    delete_api_key,
    delete_service,
    delete_webhook,
    grant_membership,
    import_accounts,
    list_api_keys,
    principal_from_user,
    reveal_account_password,
    revoke_api_key,
    revoke_membership,
    set_account_rotation_policy,
    set_rotation_enabled_setting,
    set_service_rotation_policy,
    test_webhook,
    update_account,
    update_account_registered,
    update_account_status,
    update_field_value,
    update_service_preferences,
    update_webhook,
    upsert_field_values,
)
from service_manager.webhooks import (
    list_webhook_configs,
    webhook_event_types,
)

routes = Blueprint("routes", __name__)


STATUS_ORDER = {"ativo": 0, "nunca": 1, "inativo": 2}
STATUS_LABELS = {"ativo": "Ativo", "nunca": "Nunca teve", "inativo": "Teve, mas inativo"}
OK_MESSAGES = {
    "account_added": "Conta adicionada.",
    "account_updated": "Conta atualizada.",
    "account_deleted": "Conta excluída.",
    "status_updated": "Status atualizado.",
    "registered_updated": "Cadastro atualizado.",
    "field_added": "Campo adicionado.",
    "field_saved": "Campo salvo.",
    "field_deleted": "Campo excluído.",
    "service_added": "Serviço criado.",
    "service_deleted": "Serviço excluído.",
    "bulk_updated": "Contas atualizadas.",
    "bulk_deleted": "Contas excluídas.",
    "bulk_field_created": "Campo adicionado às contas selecionadas.",
    "rotation_policy_updated": "Política de rotação atualizada.",
    "rotation_completed": "Rotação concluída.",
    "rotation_incomplete": "Rotação marcada como pendente.",
    "settings_updated": "Configurações atualizadas.",
    "api_key_deleted": "API key excluída.",
}
ROTATION_LABELS = {
    "unknown": "Desconhecido",
    "no_policy": "Sem política",
    "current": "Em dia",
    "due_soon": "Vence em breve",
    "overdue": "Vencida",
}
TEMPLATE_ROWS = [
    ("email", "password", "status"),
    ("exemplo1@gmail.com", "SenhaSegura1", "nunca"),
    ("exemplo2@gmail.com", "SenhaSegura2", "ativo"),
]

_COVERAGE_PAGE_SIZE = 100
_STATUS_RANK_SQL = "CASE link.status WHEN 'ativo' THEN 0 WHEN 'nunca' THEN 1 WHEN 'inativo' THEN 2 ELSE 1 END"
_ACCOUNT_SELECT = (
    "SELECT a.id AS id, a.email AS email, a.password_changed_at AS password_changed_at, "
    "link.status AS status, link.registered AS registered, "
    "link.rotation_days AS rotation_days, link.rotation_due_at AS rotation_due_at "
    "FROM account_service AS link JOIN accounts AS a ON a.id = link.account_id"
)
# Mirrors _rotation_state() in SQL so global counts and the rotation filter stay
# bounded. Binds :rot_today (ISO date) and :rot_sdays (service default days).
_ROTATION_STATE_SQL = """
CASE
  WHEN link.rotation_due_at IS NOT NULL THEN
    CASE
      WHEN date(link.rotation_due_at) IS NULL THEN 'unknown'
      WHEN CAST(julianday(date(link.rotation_due_at)) - julianday(date(:rot_today)) AS INTEGER) < 0 THEN 'overdue'
      WHEN CAST(julianday(date(link.rotation_due_at)) - julianday(date(:rot_today)) AS INTEGER) <= 7 THEN 'due_soon'
      ELSE 'current'
    END
  WHEN a.password_changed_at IS NULL THEN 'unknown'
  WHEN NOT (a.password_changed_at LIKE '%Z' OR a.password_changed_at LIKE '%+__:__' OR a.password_changed_at LIKE '%-__:__') THEN 'unknown'
  WHEN date(a.password_changed_at) IS NULL THEN 'unknown'
  WHEN COALESCE(link.rotation_days, :rot_sdays) IS NULL THEN 'no_policy'
  ELSE
    CASE
      WHEN CAST(julianday(date(a.password_changed_at, '+' || COALESCE(link.rotation_days, :rot_sdays) || ' days')) - julianday(date(:rot_today)) AS INTEGER) < 0 THEN 'overdue'
      WHEN CAST(julianday(date(a.password_changed_at, '+' || COALESCE(link.rotation_days, :rot_sdays) || ' days')) - julianday(date(:rot_today)) AS INTEGER) <= 7 THEN 'due_soon'
      ELSE 'current'
    END
END
"""


def _like_escape(value: str) -> str:
    return value.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


def _cursor_serializer(salt: str) -> URLSafeSerializer:
    return URLSafeSerializer(current_app.config["SECRET_KEY"], salt=salt)


def _encode_cursor(salt: str, payload: dict[str, object]) -> str:
    return _cursor_serializer(salt).dumps(payload)


def _decode_cursor(salt: str, token: str) -> dict[str, object]:
    try:
        payload = _cursor_serializer(salt).loads(token)
    except BadSignature:
        abort(400)
    if not isinstance(payload, dict):
        abort(400)
    return payload


def _account_filter_sql(*, service_id, q, status, registered, rot_state, service_days, today) -> tuple[str, dict[str, object]]:
    """Build the shared account WHERE for both the page and the aggregate queries."""
    clauses = ["link.service_id = :sid"]
    params: dict[str, object] = {"sid": service_id}
    if q:
        clauses.append("a.email LIKE :q ESCAPE '\\'")
        params["q"] = f"%{_like_escape(q)}%"
    if status:
        clauses.append("link.status = :st")
        params["st"] = status
    if registered in ("0", "1"):
        clauses.append("link.registered = :reg")
        params["reg"] = int(registered)
    if rot_state:
        clauses.append(f"({_ROTATION_STATE_SQL}) = :rot")
        params["rot"] = rot_state
        params["rot_today"] = today.isoformat()
        params["rot_sdays"] = service_days
    return " AND ".join(clauses), params


def _account_order_sql(sort: str, sql_dir: str) -> str:
    if sort == "status":
        return f"{_STATUS_RANK_SQL} {sql_dir}, a.email COLLATE NOCASE {sql_dir}, a.id {sql_dir}"
    return f"a.email COLLATE NOCASE {sql_dir}, a.id {sql_dir}"


def normalize_status(value: str | None) -> str | None:
    status = (value or "").strip().lower()
    return status if status in STATUS_ORDER else None


def _valid_email(value: str | None) -> str | None:
    return normalize_email(value) or None


def _valid_name(value: str | None) -> str | None:
    candidate = (value or "").strip()
    return candidate if 1 <= len(candidate) <= 100 else None


def _valid_secret(value: str | None) -> str | None:
    return value if isinstance(value, str) and len(value) <= 4096 else None


def _webhook_resolver():
    """Return the DNS resolver for webhook URL validation.

    Production always uses the real resolver. A test-only injected resolver is
    honored solely when the app is in testing mode, so the SSRF-bypass seam is
    never reachable from production env/config.
    """
    if current_app.testing:
        override = current_app.config.get("WEBHOOK_RESOLVER")
        if override is not None:
            return override
    return socket.getaddrinfo


def _audit_actor() -> int | None:
    user = getattr(g, "current_user", None)
    return user["id"] if user is not None else None


def _principal():
    return principal_from_user(g.current_user)


def _audit(conn, *, action: str, target_type: str, target_id: int | str | None = None, metadata: Mapping[str, object] | None = None) -> int:
    return append_audit_event(conn, action=action, target_type=target_type, target_id=target_id, actor_user_id=_audit_actor(), metadata=metadata)


def _form_error(message: str, *, status: int = 400) -> Response:
    if request.headers.get("Accept", "").startswith("application/json"):
        response = jsonify(error=message)
        response.status_code = status
        return response
    return Response(render_template("form_error.html", message=message), status=status, mimetype="text/html")


def selected_service_id(services: list, initial_service_id: int | None) -> int | None:
    raw_candidate = request.values.get("service") or request.values.get("service_id")
    if raw_candidate is None:
        if initial_service_id is not None and any(service["id"] == initial_service_id for service in services):
            return initial_service_id
        return services[0]["id"] if services else None
    try:
        candidate = int(raw_candidate)
    except (TypeError, ValueError):
        abort(404)
    if candidate <= 0 or not any(service["id"] == candidate for service in services):
        abort(404)
    return candidate


def required_service_id() -> int:
    raw_candidate = request.form.get("service_id") or request.form.get("service")
    if raw_candidate is None:
        abort(400)
    try:
        service_id = int(raw_candidate)
    except (TypeError, ValueError):
        abort(400)
    if service_id <= 0 or get_db().execute("SELECT 1 FROM services WHERE id=?", (service_id,)).fetchone() is None:
        abort(404)
    return service_id

def required_query_service_id() -> int:
    raw_candidate = request.args.get("service")
    if raw_candidate is None:
        abort(400)
    try:
        service_id = int(raw_candidate)
    except (TypeError, ValueError):
        abort(400)
    if service_id <= 0:
        abort(400)
    if get_db().execute("SELECT 1 FROM services WHERE id=?", (service_id,)).fetchone() is None:
        abort(404)
    return service_id


def _sanitize_cell(value: str) -> str:
    return "'" + value if value[:1] in ("=", "+", "-", "@") else value


_EXPORT_LIMIT = 10000


def _safe_filename_slug(name: str) -> str:
    folded = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii").lower()
    slug = re.sub(r"[^a-z0-9]+", "_", folded).strip("_")[:50].strip("_")
    return slug or "servico"


def _service_name(conn, service_id: int) -> str:
    row = conn.execute("SELECT name FROM services WHERE id=?", (service_id,)).fetchone()
    return row["name"] if row is not None else ""


def _export_row_count(conn, service_id: int) -> int:
    return conn.execute(
        "SELECT COUNT(*) AS n FROM account_service WHERE service_id = ?", (service_id,)
    ).fetchone()["n"]


def _export_fields(conn: sqlite3.Connection, service_id: int) -> tuple[tuple[int, str], ...]:
    return tuple(
        (row["id"], row["name"])
        for row in conn.execute(
            "SELECT id, name FROM custom_fields WHERE service_id = ? ORDER BY name", (service_id,)
        )
    )


def _iter_export_rows(
    conn: sqlite3.Connection, service_id: int, fields: tuple[tuple[int, str], ...]
) -> Iterator[tuple[str, ...]]:
    field_ids = [field_id for field_id, _ in fields]
    values = iter(())
    if fields:
        values = conn.execute(
            """
            SELECT a.id AS account_id, value.field_id,
                   value.value_ciphertext, value.value_nonce, value.value_key_version
            FROM field_values AS value
            JOIN custom_fields AS field ON field.id = value.field_id
            JOIN account_service AS link ON link.account_id = value.account_id
            JOIN accounts AS a ON a.id = value.account_id
            WHERE field.service_id = ? AND link.service_id = ?
            ORDER BY a.email COLLATE NOCASE, field.name
            """,
            (service_id, service_id),
        )
    pending = next(values, None)
    for account in conn.execute(
        """
        SELECT a.id AS account_id, a.email,
               a.password_ciphertext, a.password_nonce, a.password_key_version,
               link.status, link.registered
        FROM account_service AS link
        JOIN accounts AS a ON a.id = link.account_id
        WHERE link.service_id = ?
        ORDER BY a.email COLLATE NOCASE
        """,
        (service_id,),
    ):
        account_id = account["account_id"]
        current: dict[int, str] = {}
        while pending is not None and pending["account_id"] == account_id:
            current[pending["field_id"]] = decrypt_secret(
                EncryptedValue(pending["value_ciphertext"], pending["value_nonce"], pending["value_key_version"]),
                aad=account_field_aad(account_id, pending["field_id"]),
            )
            pending = next(values, None)
        password = decrypt_secret(
            EncryptedValue(account["password_ciphertext"], account["password_nonce"], account["password_key_version"]),
            aad=account_password_aad(account_id),
        )
        yield (
            _sanitize_cell(account["email"]),
            _sanitize_cell(password),
            _sanitize_cell(account["status"]),
            "sim" if account["registered"] else "não",
            *(_sanitize_cell(current.get(field_id, "")) for field_id in field_ids),
        )


_ROTATION_MAX_DAYS = 3650


def rotation_enabled(conn) -> bool:
    """Whether the global credential-rotation control is enabled. Defaults to disabled."""
    row = conn.execute("SELECT value FROM app_settings WHERE key='rotation_enabled'").fetchone()
    return row is not None and row["value"] == "1"


def require_rotation_enabled(conn) -> None:
    """Abort 404 when the rotation feature is globally disabled."""
    if not rotation_enabled(conn):
        abort(404)


def set_rotation_enabled(conn, enabled: bool) -> None:
    """Persist the global rotation flag. Caller owns the transaction."""
    conn.execute(
        "INSERT INTO app_settings (key, value) VALUES ('rotation_enabled', ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        ("1" if enabled else "0",),
    )


def _parse_rotation_days(value: str | None) -> tuple[bool, int | None]:
    """(True, None) for absent/all-whitespace (inherit/clear); (True, days) for exact ASCII decimal 1..3650; (False, None) otherwise."""
    if value is None or not value.strip():
        return (True, None)
    if not re.fullmatch(r"[0-9]+", value):
        return (False, None)
    days = int(value)
    if 1 <= days <= _ROTATION_MAX_DAYS:
        return (True, days)
    return (False, None)


def _parse_rotation_due_at(value: str | None) -> tuple[bool, str | None]:
    """(True, None) for absent/all-whitespace; (True, canonical YYYY-MM-DD) only when input exactly equals the canonical ISO date; (False, None) otherwise."""
    if value is None or not value.strip():
        return (True, None)
    try:
        canonical = date.fromisoformat(value).isoformat()
    except (ValueError, TypeError):
        return (False, None)
    if canonical != value:
        return (False, None)
    return (True, canonical)


def _rotation_state(
    password_changed_at: str | None,
    account_days: int | None,
    account_due_at: str | None,
    service_days: int | None,
    *,
    today: date | None = None,
) -> dict[str, object]:
    """Compute rotation state for one account_service link. Fails closed to 'unknown'."""
    today_utc = today or datetime.now(UTC).date()
    effective_days = account_days if account_days is not None else service_days

    def _result(state: str, due_at: str | None, days_remaining: int | None) -> dict[str, object]:
        return {"state": state, "effective_days": effective_days, "due_at": due_at, "days_remaining": days_remaining}

    # Explicit due-date override wins even when password history is unknown.
    if account_due_at is not None:
        try:
            due_date = date.fromisoformat(account_due_at)
        except (ValueError, TypeError):
            return _result("unknown", None, None)
        days_remaining = (due_date - today_utc).days
        return _result(_due_state(days_remaining), due_date.isoformat(), days_remaining)
    # No explicit override: parse the password timestamp (must be tz-aware).
    if password_changed_at is None:
        return _result("unknown", None, None)
    try:
        changed = datetime.fromisoformat(password_changed_at)
    except (ValueError, TypeError):
        return _result("unknown", None, None)
    if changed.tzinfo is None:
        return _result("unknown", None, None)
    if effective_days is None:
        return _result("no_policy", None, None)
    changed_utc = changed.astimezone(UTC).date()
    due_date = changed_utc + timedelta(days=effective_days)
    days_remaining = (due_date - today_utc).days
    return _result(_due_state(days_remaining), due_date.isoformat(), days_remaining)


def _due_state(days_remaining: int) -> str:
    if days_remaining < 0:
        return "overdue"
    if days_remaining <= 7:
        return "due_soon"
    return "current"

def _actor_label(username: str | None, metadata_json: str | None) -> str:
    if username:
        return username
    if metadata_json:
        try:
            import json
            metadata = json.loads(metadata_json)
        except (TypeError, ValueError):
            metadata = None
        if isinstance(metadata, dict):
            name = metadata.get("api_key_name")
            if isinstance(name, str) and name:
                return f"api:{name}"
    return "sistema"


def _audit_query_filters() -> tuple[str, list[object], dict[str, str]]:
    filters = {name: (request.args.get(name) or "").strip() for name in ("action", "target_type", "actor", "api_key", "since", "until", "source_ip")}
    clauses: list[str] = []
    params: list[object] = []
    if filters["action"]:
        clauses.append("e.action = ?")
        params.append(filters["action"])
    if filters["target_type"]:
        clauses.append("e.target_type = ?")
        params.append(filters["target_type"])
    if filters["actor"]:
        try:
            actor = int(filters["actor"])
        except ValueError:
            actor = 0
        if actor > 0:
            clauses.append("e.actor_user_id = ?")
            params.append(actor)
        else:
            filters["actor"] = ""
    if filters["api_key"]:
        try:
            api_key_id = int(filters["api_key"])
        except ValueError:
            api_key_id = 0
        if api_key_id > 0:
            clauses.append("e.metadata_json LIKE ?")
            params.append(f'%"api_key_id": {api_key_id}%')
        else:
            filters["api_key"] = ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", filters["since"]):
        try:
            since = date.fromisoformat(filters["since"])
        except ValueError:
            filters["since"] = ""
        else:
            clauses.append("e.occurred_at >= ?")
            params.append(since.isoformat())
    else:
        filters["since"] = ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", filters["until"]):
        try:
            until = date.fromisoformat(filters["until"])
        except ValueError:
            filters["until"] = ""
        else:
            clauses.append("e.occurred_at < ?")
            params.append((until + timedelta(days=1)).isoformat())
    else:
        filters["until"] = ""
    if filters["source_ip"]:
        escaped = filters["source_ip"].replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        clauses.append("e.source_ip LIKE ? ESCAPE '\\'")
        params.append(f"%{escaped}%")
    return (" WHERE " + " AND ".join(clauses) if clauses else "", params, filters)


@routes.get("/admin/audit")
@require_role("admin")
def audit_view() -> ResponseReturnValue:
    require_recent_reauth()
    conn = get_db()
    try:
        page = max(1, int(request.args.get("page", "1")))
    except ValueError:
        page = 1
    where, params, filters = _audit_query_filters()
    rows = conn.execute(
        f"""
        SELECT e.id, e.occurred_at, e.action, e.target_type, e.target_id, e.metadata_json, e.source_ip, u.username
        FROM audit_events AS e
        LEFT JOIN users AS u ON u.id = e.actor_user_id
        {where}
        ORDER BY e.id DESC
        LIMIT 51 OFFSET ?
        """,
        (*params, (page - 1) * 50),
    ).fetchall()
    events = [
        {
            "id": row["id"],
            "occurred_at": row["occurred_at"],
            "action": row["action"],
            "target_type": row["target_type"],
            "target_id": row["target_id"],
            "metadata_json": row["metadata_json"],
            "source_ip": row["source_ip"],
            "username": row["username"],
            "actor_label": _actor_label(row["username"], row["metadata_json"]),
        }
        for row in rows[:50]
    ]
    return Response(
        render_template("audit.html", events=events, page=page, has_next=len(rows) > 50, filters=filters, chain_healthy=verify_audit_chain(conn)),
        headers={"Cache-Control": "no-store, private"},
    )


@routes.get("/admin/audit.csv")
@require_role("admin")
def audit_csv() -> ResponseReturnValue:
    require_recent_reauth()
    conn = get_db()
    where, params, _ = _audit_query_filters()
    cursor = conn.execute(
        f"""
        SELECT e.id, e.occurred_at, u.username, e.action, e.target_type, e.target_id, e.metadata_json, e.source_ip, e.previous_hash, e.event_hash
        FROM audit_events AS e
        LEFT JOIN users AS u ON u.id = e.actor_user_id
        {where}
        ORDER BY e.id DESC
        LIMIT 10000
        """,
        params,
    )
    filename = f"auditoria_{datetime.now(UTC).strftime('%Y%m%dT%H%M%SZ')}.csv"
    spool = tempfile.SpooledTemporaryFile(max_size=8 * 1024 * 1024)
    try:
        buffer = io.StringIO(newline="")
        writer = csv.writer(buffer)

        def _flush() -> None:
            spool.write(buffer.getvalue().encode("utf-8"))
            buffer.seek(0)
            buffer.truncate(0)

        spool.write("\ufeff".encode("utf-8"))
        writer.writerow(("id", "occurred_at", "usuario", "action", "target_type", "target_id", "metadata_json", "source_ip", "previous_hash", "event_hash"))
        _flush()
        for row in cursor:
            values = list(row)
            values[2] = _actor_label(row["username"], row["metadata_json"])
            # Hash BLOBs must export as 64 lowercase hex chars, never Python bytes repr.
            values[8] = row["previous_hash"].hex() if row["previous_hash"] is not None else ""
            values[9] = row["event_hash"].hex() if row["event_hash"] is not None else ""
            writer.writerow(tuple(_sanitize_cell("" if value is None else str(value)) for value in values))
            _flush()
        spool.seek(0)
        response = send_file(spool, mimetype="text/csv", as_attachment=True, download_name=filename)
        response.call_on_close(spool.close)
        return response
    except Exception:
        spool.close()
        raise



@routes.get("/coverage")
def coverage() -> ResponseReturnValue:
    conn = get_db()
    user = g.current_user
    services = accessible_services(conn, user)
    no_access = user["role"] != "admin" and not services
    filter_mode = request.args.get("filter") or ""
    if filter_mode not in ("none-registered", "multi-active", "missing-registration"):
        filter_mode = ""
    accessible_ids = [s["id"] for s in services]
    accessible_set = set(accessible_ids)
    # Selected services matter only for missing-registration and are constrained
    # to accessible services so inaccessible IDs never widen the scan or reveal
    # that a service exists.
    selected_services = sorted(
        value
        for value in {int(raw) for raw in request.args.getlist("services") if _ASCII_SERVICE_ID.fullmatch(raw)}
        if value in accessible_set
    )
    # Selected services are meaningful only for missing-registration; drop them
    # elsewhere so they never leak into query state or pagination links.
    if filter_mode != "missing-registration":
        selected_services = []
    accounts: list[Any] = []
    aggregates: dict[int, dict[str, int]] = {}
    links_by_account: dict[tuple[int, int], dict[str, object]] = {}
    total_count = 0
    page_count = 0
    has_prev = has_next = False
    prev_cursor: str | None = None
    next_cursor: str | None = None
    view_params: dict[str, object] = {}
    if filter_mode:
        view_params["filter"] = filter_mode
    if selected_services:
        view_params["services"] = selected_services

    if services:
        svc_ph = ",".join("?" for _ in accessible_ids)
        # Admins see every account (including unlinked ones); members see only
        # accounts linked to a service they can access.
        join = "LEFT JOIN" if user["role"] == "admin" else "JOIN"
        base_from = (
            f"FROM accounts AS a {join} account_service AS link "
            f"ON link.account_id = a.id AND link.service_id IN ({svc_ph})"
        )
        having = ""
        having_params: list[object] = []
        if filter_mode == "none-registered":
            having = "HAVING registered_count = 0"
        elif filter_mode == "multi-active":
            having = "HAVING active_count > 1"
        elif filter_mode == "missing-registration" and selected_services:
            sel_ph = ",".join("?" for _ in selected_services)
            having = (
                "HAVING COALESCE(SUM(CASE WHEN link.service_id IN "
                f"({sel_ph}) THEN link.registered ELSE 0 END), 0) < ?"
            )
            having_params = [*selected_services, len(selected_services)]
        agg_cols = (
            "COALESCE(SUM(link.registered), 0) AS registered_count, "
            "COALESCE(SUM(CASE WHEN link.status = 'ativo' THEN 1 ELSE 0 END), 0) AS active_count"
        )
        total_count = conn.execute(
            f"SELECT COUNT(*) AS n FROM (SELECT a.id, {agg_cols} {base_from} GROUP BY a.id {having})",
            (*accessible_ids, *having_params),
        ).fetchone()["n"]

        limit = _COVERAGE_PAGE_SIZE + 1
        keyset = ""
        key_params: list[object] = []
        nav = "next"
        order_dir = "ASC"
        cursor_token = request.args.get("cursor")
        if cursor_token:
            cur = _decode_cursor("coverage-cursor", cursor_token)
            if not isinstance(cur.get("e"), str) or not isinstance(cur.get("id"), int) or cur.get("nav") not in ("next", "prev"):
                abort(400)
            # The cursor is bound to the filter/services it was issued under, so a
            # tampered cursor replayed under a different filter is rejected rather
            # than paginating from the wrong anchor.
            if cur.get("f") != filter_mode or cur.get("sv") != selected_services:
                abort(400)
            nav = cur["nav"]
            comparator = ">" if nav == "next" else "<"
            order_dir = "ASC" if nav == "next" else "DESC"
            keyset = (
                f"AND (a.email COLLATE NOCASE {comparator} ? "
                f"OR (a.email COLLATE NOCASE = ? AND a.id {comparator} ?))"
            )
            key_params = [cur["e"], cur["e"], cur["id"]]
        rows_sql = (
            f"SELECT a.id AS id, a.email AS email, {agg_cols} {base_from} "
            f"WHERE 1=1 {keyset} GROUP BY a.id, a.email {having} "
            f"ORDER BY a.email COLLATE NOCASE {order_dir}, a.id {order_dir} LIMIT ?"
        )
        fetched = conn.execute(rows_sql, (*accessible_ids, *key_params, *having_params, limit)).fetchall()
        has_more = len(fetched) > _COVERAGE_PAGE_SIZE
        page = fetched[:_COVERAGE_PAGE_SIZE]
        if nav == "prev":
            page = list(reversed(page))
            has_prev = has_more
            has_next = True
        else:
            has_next = has_more
            has_prev = bool(cursor_token)
        accounts = page
        page_count = len(page)
        aggregates = {
            row["id"]: {"registered_count": row["registered_count"], "active_count": row["active_count"]}
            for row in page
        }
        if page:
            if has_prev:
                prev_cursor = _encode_cursor("coverage-cursor", {"e": page[0]["email"], "id": page[0]["id"], "nav": "prev", "f": filter_mode, "sv": selected_services})
            if has_next:
                next_cursor = _encode_cursor("coverage-cursor", {"e": page[-1]["email"], "id": page[-1]["id"], "nav": "next", "f": filter_mode, "sv": selected_services})
            page_ids = [row["id"] for row in page]
            id_ph = ",".join("?" for _ in page_ids)
            # Only queried links are materialized; absent pairs stay out of the map
            # and the template renders the default cell for a missing lookup.
            for lr in conn.execute(
                f"SELECT account_id, service_id, status, registered FROM account_service "
                f"WHERE account_id IN ({id_ph}) AND service_id IN ({svc_ph})",
                (*page_ids, *accessible_ids),
            ):
                links_by_account[(lr["account_id"], lr["service_id"])] = {
                    "status": lr["status"],
                    "registered": bool(lr["registered"]),
                }
    return Response(
        render_template(
            "coverage.html", services=services, accounts=accounts, aggregates=aggregates,
            links_by_account=links_by_account, labels=STATUS_LABELS, no_access=no_access,
            filter_mode=filter_mode, selected_services=selected_services, total_count=total_count,
            page_count=page_count, has_prev=has_prev, has_next=has_next,
            prev_cursor=prev_cursor, next_cursor=next_cursor, view_params=view_params,
        ),
        headers={"Cache-Control": "no-store, private"},
    )


def encrypted_account_password(account_id: int, password: str) -> tuple[bytes, bytes, int]:
    value = encrypt_secret(password, aad=account_password_aad(account_id))
    return value.ciphertext, value.nonce, value.key_version


def encrypted_field_value(account_id: int, field_id: int, value: str) -> tuple[bytes, bytes, int]:
    encrypted = encrypt_secret(value, aad=account_field_aad(account_id, field_id))
    return encrypted.ciphertext, encrypted.nonce, encrypted.key_version




def _related_account(conn, service_id: int | None, account_id: int) -> None:
    if service_id is None or conn.execute(
        "SELECT 1 FROM account_service WHERE account_id=? AND service_id=?", (account_id, service_id)
    ).fetchone() is None:
        abort(404)


def _related_field_account(conn, service_id: int | None, field_id: int, account_id: int) -> None:
    _related_account(conn, service_id, account_id)
    if conn.execute("SELECT 1 FROM custom_fields WHERE id=? AND service_id=?", (field_id, service_id)).fetchone() is None:
        abort(404)


def link_all_services(conn, account_id: int, active_service_id: int, status: str, registered: int = 0) -> None:
    conn.execute(
        """
        INSERT INTO account_service (account_id, service_id, status, registered)
        SELECT ?, id, CASE WHEN id = ? THEN ? ELSE 'nunca' END, CASE WHEN id = ? THEN ? ELSE 0 END
        FROM services WHERE true
        ON CONFLICT(account_id, service_id) DO UPDATE SET status = excluded.status, registered = excluded.registered
        """,
        (account_id, active_service_id, status, active_service_id, registered),
    )

@routes.get("/healthz")
def healthz() -> ResponseReturnValue:
    try:
        conn = get_db()
        healthy = conn.execute("SELECT 1").fetchone() is not None and schema_is_current(conn) and verify_audit_chain(conn)
    except (sqlite3.Error, OSError, RuntimeError):
        healthy = False
    if not healthy:
        return jsonify(status="degraded"), 503
    return jsonify(status="ok")


class _InvalidServicePreferences(ValueError):
    pass


_ASCII_SERVICE_ID = re.compile(r"[1-9][0-9]*\Z", re.ASCII)


@routes.post("/preferences/services")
def service_preferences_update() -> Response:
    raw_service_ids = request.form.getlist("service_ids")
    initial_values = request.form.getlist("initial_service_id")
    if any(_ASCII_SERVICE_ID.fullmatch(value) is None for value in (*raw_service_ids, *initial_values)):
        return Response("Preferências de serviços inválidas", status=400)
    service_ids = [int(value) for value in raw_service_ids]
    if len(service_ids) != len(set(service_ids)):
        return Response("Preferências de serviços inválidas", status=400)
    initial_service_id = int(initial_values[0]) if len(initial_values) == 1 else None
    conn = get_db()
    try:
        with transaction(conn):
            services = accessible_services(conn, g.current_user)
            accessible_ids = [service["id"] for service in services]
            if len(initial_values) != (1 if accessible_ids else 0):
                raise _InvalidServicePreferences
            if sorted(service_ids) != sorted(accessible_ids):
                raise _InvalidServicePreferences
            if initial_service_id is not None and initial_service_id not in service_ids:
                raise _InvalidServicePreferences
            update_service_preferences(
                conn,
                _principal(),
                user_id=g.current_user["id"],
                service_ids=service_ids,
                initial_service_id=initial_service_id,
            )
    except _InvalidServicePreferences:
        return Response("Preferências de serviços inválidas", status=400)
    return Response(status=204)


@routes.get("/")
def index() -> str:
    conn = get_db()
    user = g.current_user
    services = accessible_services(conn, user)
    initial_service_id = next((service["id"] for service in services if service["is_initial"]), None)
    raw_service = request.values.get("service") or request.values.get("service_id")
    if raw_service is not None and not any(str(s["id"]) == raw_service for s in services):
        # A syntactically valid, existing service the caller cannot access -> 403, never fallback.
        try:
            candidate = int(raw_service)
        except (TypeError, ValueError):
            abort(404)
        if candidate > 0 and conn.execute("SELECT 1 FROM services WHERE id=?", (candidate,)).fetchone() is not None:
            abort(403)
        abort(404)
    service_id = selected_service_id(services, initial_service_id)
    service_role = get_user_service_role(conn, user, service_id) if service_id is not None else None
    no_access = user["role"] != "admin" and not services
    can_reveal = service_role in ("admin", "editor", "service_admin")
    can_edit = can_reveal
    can_import = service_role in ("admin", "service_admin")
    can_export = can_import
    can_delete = can_import
    capabilities = {
        "reveal": can_reveal,
        "edit": can_edit,
        "import": can_import,
        "export": can_export,
        "delete": can_delete,
    }
    rot_enabled = rotation_enabled(conn)
    today = datetime.now(UTC).date()
    rot_filter = request.args.get("rot") or ""
    if not rot_enabled or rot_filter not in ("due_soon", "overdue", "unknown", "current", "no_policy"):
        rot_filter = ""
    sort = request.args.get("sort")
    if sort not in ("email", "status"):
        sort = "status"
    direction = request.args.get("dir")
    if direction not in ("asc", "desc"):
        direction = "asc"
    ascending = direction == "asc"
    q = (request.args.get("q") or "").strip()
    st_filter = request.args.get("st") or ""
    if st_filter not in STATUS_ORDER:
        st_filter = ""
    reg_filter = request.args.get("reg") or ""
    if reg_filter not in ("0", "1"):
        reg_filter = ""
    filters_active = bool(q or st_filter or reg_filter or rot_filter)

    rows: list[dict[str, object]] = []
    fields: list[Any] = []
    counts = {status: 0 for status in STATUS_ORDER}
    rotation_counts = {"due_soon": 0, "overdue": 0}
    service_days = None
    page_count = 0

    if service_id is not None:
        if rot_enabled:
            service_days_row = conn.execute("SELECT rotation_days FROM services WHERE id=?", (service_id,)).fetchone()
            service_days = service_days_row["rotation_days"] if service_days_row is not None else None
        fields = conn.execute("SELECT id, name FROM custom_fields WHERE service_id = ? ORDER BY name", (service_id,)).fetchall()

        filter_sql, base_params = _account_filter_sql(
            service_id=service_id, q=q, status=st_filter, registered=reg_filter,
            rot_state=rot_filter, service_days=service_days, today=today,
        )
        sql = f"{_ACCOUNT_SELECT} WHERE {filter_sql} ORDER BY {_account_order_sql(sort, 'ASC' if ascending else 'DESC')}"
        page = conn.execute(sql, base_params).fetchall()

        for row in page:
            entry: dict[str, object] = {
                "id": row["id"],
                "email": row["email"],
                "status": row["status"],
                "registered": bool(row["registered"]),
            }
            if rot_enabled:
                rotation = _rotation_state(row["password_changed_at"], row["rotation_days"], row["rotation_due_at"], service_days, today=today)
                entry.update({
                    "rotation_state": rotation["state"],
                    "rotation_due_at": rotation["due_at"],
                    "rotation_days_remaining": rotation["days_remaining"],
                    "rotation_effective_days": rotation["effective_days"],
                    "rotation_days_link": row["rotation_days"],
                    "rotation_due_at_link": row["rotation_due_at"],
                })
            rows.append(entry)
        page_count = len(rows)

        for crow in conn.execute("SELECT link.status AS s, COUNT(*) AS n FROM account_service AS link WHERE link.service_id = ? GROUP BY link.status", (service_id,)):
            if crow["s"] in counts:
                counts[crow["s"]] = crow["n"]
        if rot_enabled:
            for rrow in conn.execute(
                f"SELECT ({_ROTATION_STATE_SQL}) AS s, COUNT(*) AS n FROM account_service AS link "
                f"JOIN accounts AS a ON a.id = link.account_id WHERE link.service_id = :sid GROUP BY s",
                {"sid": service_id, "rot_today": today.isoformat(), "rot_sdays": service_days},
            ):
                if rrow["s"] in rotation_counts:
                    rotation_counts[rrow["s"]] = rrow["n"]
    counts["total"] = sum(counts[status] for status in STATUS_ORDER)
    current_name = next((service["name"] for service in services if service["id"] == service_id), None)
    feedback = None
    error_kind = request.args.get("error")
    error_messages = {
        "format": "formato inválido",
        "limits": "limites excedidos",
        "validation": "dados inválidos",
    }
    feedback_is_error = error_kind in error_messages
    if feedback_is_error:
        feedback = f"Importação rejeitada: {error_messages[error_kind]}. 0 adicionadas; 0 ignoradas."
    elif request.args.get("added") is not None or request.args.get("skipped") is not None:
        try:
            added = max(0, int(request.args.get("added", "0")))
            skipped = max(0, int(request.args.get("skipped", "0")))
        except ValueError:
            abort(400)
        feedback = f"Importação concluída: {added} adicionadas; {skipped} ignoradas."
    elif (ok := request.args.get("ok")) in OK_MESSAGES:
        feedback = OK_MESSAGES[ok]
    effective_initial_service_id = initial_service_id or (services[0]["id"] if services else None)
    view_params: dict[str, object] = {"sort": sort, "dir": direction}
    if q:
        view_params["q"] = q
    if st_filter:
        view_params["st"] = st_filter
    if reg_filter:
        view_params["reg"] = reg_filter
    if rot_filter:
        view_params["rot"] = rot_filter
    return render_template(
        "index.html", rows=rows, labels=STATUS_LABELS, counts=counts, services=services, current=service_id,
        current_name=current_name, initial_service_id=effective_initial_service_id, service_fields=fields,
        feedback=feedback, feedback_is_error=feedback_is_error, service_role=service_role, capabilities=capabilities,
        no_access=no_access, rotation_counts=rotation_counts, rot_filter=rot_filter, rotation_labels=ROTATION_LABELS,
        service_rotation_days=(service_days if service_id is not None else None), rotation_enabled=rot_enabled,
        q=q, st_filter=st_filter, reg_filter=reg_filter, sort=sort, direction=direction, filters_active=filters_active,
        page_count=page_count, view_params=view_params,
    )


@routes.get("/accounts/<int:account_id>/details")
def account_details(account_id: int) -> ResponseReturnValue:
    conn = get_db()
    service_id = required_query_service_id()
    require_account_role(conn, account_id, service_id, "viewer")
    user = g.current_user
    service_role = get_user_service_role(conn, user, service_id)
    can_edit = service_role in ("admin", "editor", "service_admin")
    can_delete = service_role in ("admin", "service_admin")
    fields: list[dict[str, object]] = []
    for row in conn.execute(
        """
        SELECT value.field_id AS field_id, field.name AS name,
               value.value_ciphertext AS ct, value.value_nonce AS nonce, value.value_key_version AS kv
        FROM field_values AS value
        JOIN custom_fields AS field ON field.id = value.field_id
        WHERE field.service_id = ? AND value.account_id = ?
        ORDER BY field.name
        """,
        (service_id, account_id),
    ):
        value = decrypt_secret(
            EncryptedValue(row["ct"], row["nonce"], row["kv"]),
            aad=account_field_aad(account_id, row["field_id"]),
        )
        fields.append({"field_id": row["field_id"], "name": row["name"], "value": value})
    return Response(
        render_template(
            "_account_details.html",
            account_id=account_id, current=service_id, fields=fields,
            capabilities={"edit": can_edit, "delete": can_delete},
        ),
        headers={"Cache-Control": "no-store, private"},
    )


@routes.post("/add")
def add() -> ResponseReturnValue:
    conn = get_db()
    service_id = required_service_id()
    require_service_role(conn, service_id, "editor")
    email = _valid_email(request.form.get("email"))
    password = _valid_secret(request.form.get("password", ""))
    status = normalize_status(request.form.get("status"))
    registered = 1 if request.form.get("registered") == "1" else 0
    if email is None or password is None or status is None:
        return _form_error("Conta inválida")
    try:
        with transaction(conn):
            create_account(
                conn,
                _principal(),
                service_id=service_id,
                email=email,
                password=password,
                status=status,
                registered=registered,
            )
    except DomainError as error:
        # HTML/form contract keeps 400 for validation and uniqueness conflicts.
        return _form_error(error.message, status=400)
    return redirect(url_for("routes.index", service=service_id, ok="account_added"))


@routes.get("/template.csv")
def template_csv() -> ResponseReturnValue:
    conn = get_db()
    service_id = required_query_service_id()
    require_service_role(conn, service_id, "viewer")
    stream = io.StringIO()
    csv.writer(stream).writerows(TEMPLATE_ROWS)
    return Response(stream.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=modelo_credenciais.csv"})


@routes.get("/template.xlsx")
def template_xlsx() -> ResponseReturnValue:
    from openpyxl import Workbook

    conn = get_db()
    service_id = required_query_service_id()
    require_service_role(conn, service_id, "viewer")
    workbook = Workbook()
    worksheet = workbook.active
    if worksheet is None:
        raise RuntimeError("new workbook has no active worksheet")
    for row in TEMPLATE_ROWS:
        worksheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    return Response(stream.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=modelo_credenciais.xlsx"})

@routes.get("/export.csv")
def export_csv() -> ResponseReturnValue:
    conn = get_db()
    service_id = required_query_service_id()
    require_service_role(conn, service_id, "service_admin")
    require_recent_reauth()
    count = _export_row_count(conn, service_id)
    if count > _EXPORT_LIMIT:
        return _form_error("Exportação limitada a 10000 contas.", status=413)
    stamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    slug = _safe_filename_slug(_service_name(conn, service_id))
    filename = f"contas_{slug}_{service_id}_{stamp}.csv"
    fields = _export_fields(conn, service_id)
    headers = ("email", "password", "status", "cadastrada", *(_sanitize_cell(f"campo:{name}") for _, name in fields))
    spool = tempfile.SpooledTemporaryFile(max_size=8 * 1024 * 1024)
    try:
        buffer = io.StringIO(newline="")
        writer = csv.writer(buffer)

        def _flush() -> None:
            spool.write(buffer.getvalue().encode("utf-8"))
            buffer.seek(0)
            buffer.truncate(0)

        spool.write("\ufeff".encode("utf-8"))
        writer.writerow(headers)
        _flush()
        for row in _iter_export_rows(conn, service_id, fields):
            writer.writerow(row)
            _flush()
        spool.seek(0)
        response = send_file(spool, mimetype="text/csv", as_attachment=True, download_name=filename)
        response.call_on_close(spool.close)
        with transaction(conn):
            _audit(conn, action="accounts.exported", target_type="service", target_id=service_id, metadata={"rows": count, "format": "csv"})
        return response
    except Exception:
        spool.close()
        raise


@routes.get("/export.xlsx")
def export_xlsx() -> ResponseReturnValue:
    from openpyxl import Workbook

    conn = get_db()
    service_id = required_query_service_id()
    require_service_role(conn, service_id, "service_admin")
    require_recent_reauth()
    count = _export_row_count(conn, service_id)
    if count > _EXPORT_LIMIT:
        return _form_error("Exportação limitada a 10000 contas.", status=413)
    stamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    slug = _safe_filename_slug(_service_name(conn, service_id))
    filename = f"contas_{slug}_{service_id}_{stamp}.xlsx"
    fields = _export_fields(conn, service_id)
    headers = ("email", "password", "status", "cadastrada", *(_sanitize_cell(f"campo:{name}") for _, name in fields))
    spool = tempfile.SpooledTemporaryFile(max_size=8 * 1024 * 1024)
    try:
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet()
        worksheet.append(headers)
        for row in _iter_export_rows(conn, service_id, fields):
            worksheet.append(row)
        workbook.save(spool)
        spool.seek(0)
        response = send_file(
            spool,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
        response.call_on_close(spool.close)
        with transaction(conn):
            _audit(conn, action="accounts.exported", target_type="service", target_id=service_id, metadata={"rows": count, "format": "xlsx"})
        return response
    except Exception:
        spool.close()
        raise



@routes.post("/import")
def import_bulk() -> ResponseReturnValue:
    conn = get_db()
    service_id = required_service_id()
    require_service_role(conn, service_id, "service_admin")
    upload = request.files.get("file")
    if upload is None or not upload.filename:
        return redirect(url_for("routes.index", service=service_id, error="format"))
    if not has_allowed_upload_mimetype(upload.filename, upload.mimetype):
        return redirect(url_for("routes.index", service=service_id, error="format"))
    try:
        records = parse_import_file(upload.filename, upload.stream)
    except ImportFormatError as error:
        return redirect(url_for("routes.index", service=service_id, error=error.kind))

    normalized_records: list[tuple[str, str, str, tuple[str, ...]]] = []
    for record in records.records:
        normalized_email = _valid_email(record.email)
        normalized_status = normalize_status(record.status)
        if normalized_email is None or normalized_status is None or _valid_secret(record.password) is None:
            return redirect(url_for("routes.index", service=service_id, error="validation"))
        if any(_valid_secret(value) is None for value in record.field_values):
            return redirect(url_for("routes.index", service=service_id, error="validation"))
        normalized_records.append((normalized_email, record.password, normalized_status, record.field_values))
    field_names = records.field_names
    if any(_valid_name(name) is None for name in field_names):
        return redirect(url_for("routes.index", service=service_id, error="validation"))

    try:
        with transaction(conn):
            added, skipped = import_accounts(
                conn,
                _principal(),
                service_id=service_id,
                records=normalized_records,
                field_names=field_names,
            )
    except Exception:
        current_app.logger.exception("import transaction failed")
        return redirect(url_for("routes.index", service=service_id, error="validation"))
    return redirect(url_for("routes.index", service=service_id, added=added, skipped=skipped))


@routes.post("/update/<int:item_id>")
@routes.post("/accounts/<int:item_id>")
def update(item_id: int) -> ResponseReturnValue:
    conn = get_db()
    service_id = required_service_id()
    require_account_role(conn, item_id, service_id, "editor")
    email = _valid_email(request.form.get("email"))
    password = _valid_secret(request.form.get("password", ""))
    if email is None or password is None:
        return _form_error("Conta inválida")
    try:
        with transaction(conn):
            update_account(
                conn,
                _principal(),
                account_id=item_id,
                service_id=service_id,
                email=email,
                password=password,
            )
    except DomainError as error:
        return _form_error(error.message, status=400)
    return redirect(url_for("routes.index", service=service_id, ok="account_updated"))


@routes.post("/accounts/<int:item_id>/status")
def update_status(item_id: int) -> ResponseReturnValue:
    conn = get_db()
    service_id = required_service_id()
    require_account_role(conn, item_id, service_id, "editor")
    status = normalize_status(request.form.get("status"))
    if status is None:
        return _form_error("Status inválido")
    with transaction(conn):
        update_account_status(conn, _principal(), account_id=item_id, service_id=service_id, status=status)
    return redirect(url_for("routes.index", service=service_id, ok="status_updated"))


@routes.post("/accounts/<int:item_id>/registered")
def update_registered(item_id: int) -> ResponseReturnValue:
    conn = get_db()
    service_id = required_service_id()
    require_account_role(conn, item_id, service_id, "editor")
    raw = request.form.get("registered", "0")
    if raw not in {"0", "1"}:
        return _form_error("Cadastro inválido")
    registered = int(raw)
    with transaction(conn):
        update_account_registered(conn, _principal(), account_id=item_id, service_id=service_id, registered=registered)
    return redirect(url_for("routes.index", service=service_id, ok="registered_updated"))

@routes.post("/service/<int:service_id>/rotation-policy")
def service_rotation_policy(service_id: int) -> ResponseReturnValue:
    conn = get_db()
    require_rotation_enabled(conn)
    if conn.execute("SELECT 1 FROM services WHERE id=?", (service_id,)).fetchone() is None:
        abort(404)
    require_service_role(conn, service_id, "service_admin")
    ok, days = _parse_rotation_days(request.form.get("rotation_days"))
    if not ok:
        return _form_error("Intervalo inválido")
    with transaction(conn):
        set_service_rotation_policy(conn, _principal(), service_id=service_id, rotation_days=days)
    return redirect(url_for("routes.index", service=service_id, ok="rotation_policy_updated"))


@routes.post("/accounts/<int:account_id>/rotation-policy")
def account_rotation_policy(account_id: int) -> ResponseReturnValue:
    conn = get_db()
    require_rotation_enabled(conn)
    service_id = required_service_id()
    require_account_role(conn, account_id, service_id, "editor")
    days_ok, days = _parse_rotation_days(request.form.get("rotation_days"))
    due_ok, due_at = _parse_rotation_due_at(request.form.get("rotation_due_at"))
    if not days_ok or not due_ok:
        return _form_error("Política de rotação inválida")
    with transaction(conn):
        set_account_rotation_policy(
            conn,
            _principal(),
            account_id=account_id,
            service_id=service_id,
            rotation_days=days,
            rotation_due_at=due_at,
        )
    return redirect(url_for("routes.index", service=service_id, ok="rotation_policy_updated"))


@routes.get("/rotation")
def rotation_view() -> ResponseReturnValue:
    conn = get_db()
    require_rotation_enabled(conn)
    service_id = required_query_service_id()
    granted = require_service_role(conn, service_id, "viewer")
    service_row = conn.execute("SELECT name, rotation_days FROM services WHERE id=?", (service_id,)).fetchone()
    if service_row is None:
        abort(404)
    service_days = service_row["rotation_days"]
    can_edit = granted in ("admin", "editor", "service_admin")
    accounts = []
    today = datetime.now(UTC).date()
    for row in conn.execute(
        """
        SELECT a.id, a.email, a.password_changed_at, link.rotation_days, link.rotation_due_at
        FROM account_service AS link
        JOIN accounts AS a ON a.id = link.account_id
        WHERE link.service_id = ?
        ORDER BY a.email COLLATE NOCASE
        """,
        (service_id,),
    ):
        state = _rotation_state(row["password_changed_at"], row["rotation_days"], row["rotation_due_at"], service_days, today=today)
        if state["state"] in ("unknown", "due_soon", "overdue"):
            accounts.append({
                "id": row["id"],
                "email": row["email"],
                "password_changed_at": row["password_changed_at"],
                "effective_days": state["effective_days"],
                "due_at": state["due_at"],
                "days_remaining": state["days_remaining"],
                "state": state["state"],
            })
    return Response(
        render_template("rotation.html", accounts=accounts, service_id=service_id, current_name=service_row["name"], labels=ROTATION_LABELS, can_edit=can_edit),
        headers={"Cache-Control": "no-store, private"},
    )


@routes.post("/accounts/<int:account_id>/rotation")
def complete_rotation(account_id: int) -> ResponseReturnValue:
    conn = get_db()
    require_rotation_enabled(conn)
    service_id = required_service_id()
    require_account_role(conn, account_id, service_id, "editor")
    outcome = request.form.get("outcome", "")
    if outcome not in ("completed", "incomplete"):
        return _form_error("Resultado inválido")
    new_password = _valid_secret(request.form.get("new_password", "")) if outcome == "completed" else None
    if outcome == "completed" and not new_password:
        return _form_error("Nova senha obrigatória")
    try:
        with transaction(conn):
            op_complete_rotation(
                conn,
                _principal(),
                account_id=account_id,
                service_id=service_id,
                outcome=outcome,
                new_password=new_password,
            )
    except DomainError as error:
        return _form_error(error.message, status=400)
    ok = "rotation_incomplete" if outcome == "incomplete" else "rotation_completed"
    return redirect(url_for("routes.rotation_view", service=service_id, ok=ok))


def _bulk_account_ids() -> list[int] | Response:
    raw_ids = request.form.getlist("account_ids")
    try:
        account_ids = [int(raw) for raw in raw_ids]
    except (TypeError, ValueError):
        return _form_error("Seleção inválida")
    if any(account_id <= 0 for account_id in account_ids):
        return _form_error("Seleção inválida")
    unique = list(dict.fromkeys(account_ids))
    if not unique or len(unique) > 200:
        return _form_error("Seleção inválida")
    return unique


@routes.post("/accounts/bulk/status")
def bulk_status() -> ResponseReturnValue:
    conn = get_db()
    service_id = required_service_id()
    account_ids = _bulk_account_ids()
    if isinstance(account_ids, Response):
        return account_ids
    status = normalize_status(request.form.get("status"))
    if status is None:
        return _form_error("Status inválido")
    require_accounts_role(conn, account_ids, service_id, "editor")
    with transaction(conn):
        bulk_update_status(conn, _principal(), service_id=service_id, account_ids=account_ids, status=status)
    return redirect(url_for("routes.index", service=service_id, ok="bulk_updated"))


@routes.post("/accounts/bulk/registered")
def bulk_registered() -> ResponseReturnValue:
    conn = get_db()
    service_id = required_service_id()
    account_ids = _bulk_account_ids()
    if isinstance(account_ids, Response):
        return account_ids
    raw = request.form.get("registered", "")
    if raw not in {"0", "1"}:
        return _form_error("Cadastro inválido")
    registered = int(raw)
    require_accounts_role(conn, account_ids, service_id, "editor")
    with transaction(conn):
        bulk_update_registered(conn, _principal(), service_id=service_id, account_ids=account_ids, registered=registered)
    return redirect(url_for("routes.index", service=service_id, ok="bulk_updated"))


@routes.post("/accounts/bulk/field")
def bulk_field() -> ResponseReturnValue:
    conn = get_db()
    service_id = required_service_id()
    account_ids = _bulk_account_ids()
    if isinstance(account_ids, Response):
        return account_ids
    try:
        field_id = int(request.form.get("field_id", ""))
    except (TypeError, ValueError):
        return _form_error("Campo inválido")
    value = _valid_secret(request.form.get("field_value", ""))
    if field_id <= 0 or value is None or not value:
        return _form_error("Campo inválido")
    require_accounts_role(conn, account_ids, service_id, "editor")
    try:
        with transaction(conn):
            bulk_set_field(
                conn,
                _principal(),
                service_id=service_id,
                account_ids=account_ids,
                field_id=field_id,
                value=value,
            )
    except NotFoundError:
        abort(404)
    except DomainError as error:
        return _form_error(error.message, status=400)
    return redirect(url_for("routes.index", service=service_id, ok="bulk_updated"))


@routes.post("/accounts/bulk/field/add")
def bulk_field_add() -> ResponseReturnValue:
    conn = get_db()
    service_id = required_service_id()
    account_ids = _bulk_account_ids()
    if isinstance(account_ids, Response):
        return account_ids
    name = _valid_name(request.form.get("field_name"))
    if name is None:
        return _form_error("Campo inválido")
    require_accounts_role(conn, account_ids, service_id, "editor")
    with transaction(conn):
        bulk_create_field(conn, _principal(), service_id=service_id, account_ids=account_ids, name=name)
    return redirect(url_for("routes.index", service=service_id, ok="bulk_field_created"))


@routes.post("/accounts/bulk/delete")
def bulk_delete() -> ResponseReturnValue:
    conn = get_db()
    service_id = required_service_id()
    account_ids = _bulk_account_ids()
    if isinstance(account_ids, Response):
        return account_ids
    if request.form.get("confirmation_count", "") != str(len(account_ids)):
        return _form_error("Confirmação inválida")
    require_accounts_role(conn, account_ids, service_id, "service_admin", all_linked_services=True)
    with transaction(conn):
        bulk_delete_accounts(conn, _principal(), service_id=service_id, account_ids=account_ids)
    return redirect(url_for("routes.index", service=service_id, ok="bulk_deleted"))

@routes.post("/delete/<int:item_id>")
def delete(item_id: int) -> ResponseReturnValue:
    conn = get_db()
    service_id = required_service_id()
    require_account_role(conn, item_id, service_id, "service_admin", all_linked_services=True)
    with transaction(conn):
        delete_account(conn, _principal(), account_id=item_id, service_id=service_id)
    return redirect(url_for("routes.index", service=service_id, ok="account_deleted"))


@routes.post("/service/add")
@require_role("admin")
def service_add() -> ResponseReturnValue:
    name = _valid_name(request.form.get("name"))
    if name is None:
        return _form_error("Serviço inválido")
    conn = get_db()
    with transaction(conn):
        service_id = create_service(conn, _principal(), name)
    return redirect(url_for("routes.index", ok="service_added", service=service_id))


@routes.post("/service/delete/<int:service_id>")
@require_role("admin")
def service_delete(service_id: int) -> ResponseReturnValue:
    conn = get_db()
    try:
        with transaction(conn):
            delete_service(conn, _principal(), service_id)
    except NotFoundError:
        abort(404)
    return redirect(url_for("routes.index", ok="service_deleted"))


@routes.post("/field/add")
def field_add() -> ResponseReturnValue:
    conn = get_db()
    service_id = required_service_id()
    name = _valid_name(request.form.get("name"))
    value = _valid_secret(request.form.get("value", ""))
    raw_ids = request.form.getlist("account_ids")
    try:
        account_ids = [int(raw) for raw in raw_ids]
    except (TypeError, ValueError):
        return _form_error("Campo inválido")
    if name is None or value is None or not account_ids or any(account_id <= 0 for account_id in account_ids):
        return _form_error("Campo inválido")
    for account_id in account_ids:
        require_account_role(conn, account_id, service_id, "editor")
    try:
        with transaction(conn):
            upsert_field_values(
                conn,
                _principal(),
                service_id=service_id,
                name=name,
                account_ids=account_ids,
                value=value,
            )
    except DomainError as error:
        return _form_error(error.message, status=400)
    return redirect(url_for("routes.index", service=service_id, ok="field_added", _anchor=f"row-{account_ids[0]}"))


@routes.post("/field/update/<int:field_id>/<int:account_id>")
def field_update(field_id: int, account_id: int) -> ResponseReturnValue:
    conn = get_db()
    service_id = required_service_id()
    _related_field_account(conn, service_id, field_id, account_id)
    require_account_role(conn, account_id, service_id, "editor")
    value = _valid_secret(request.form.get("value", ""))
    if value is None:
        return _form_error("Campo inválido")
    try:
        with transaction(conn):
            update_field_value(
                conn,
                _principal(),
                service_id=service_id,
                field_id=field_id,
                account_id=account_id,
                value=value,
            )
    except NotFoundError:
        abort(404)
    except DomainError as error:
        return _form_error(error.message, status=400)
    return redirect(url_for("routes.index", service=service_id, ok="field_saved", _anchor=f"row-{account_id}"))


@routes.post("/field/delete/<int:field_id>/<int:account_id>")
def field_delete(field_id: int, account_id: int) -> ResponseReturnValue:
    conn = get_db()
    service_id = required_service_id()
    _related_field_account(conn, service_id, field_id, account_id)
    require_account_role(conn, account_id, service_id, "service_admin")
    with transaction(conn):
        delete_field_value(
            conn,
            _principal(),
            service_id=service_id,
            field_id=field_id,
            account_id=account_id,
        )
    return redirect(url_for("routes.index", service=service_id, ok="field_deleted", _anchor=f"row-{account_id}"))


@routes.post("/api/accounts/<int:account_id>/secrets/password/reveal")
def reveal_password(account_id: int) -> ResponseReturnValue:
    user = g.current_user
    conn = get_db()
    service_id = required_query_service_id()
    require_account_role(conn, account_id, service_id, "editor")
    try:
        with transaction(conn):
            value = reveal_account_password(
                conn,
                _principal(),
                account_id=account_id,
                subject=str(user["id"]),
                source_ip=source_ip(),
                actor_user_id_for_webhook=int(user["id"]),
            )
    except DomainError as error:
        if error.status == 429:
            return Response(error.message, status=429)
        if error.status == 404:
            abort(404)
        return Response(error.message, status=error.status)
    response = jsonify(value=value, expires_in=30)
    response.headers["Cache-Control"] = "no-store, private"
    return response


@routes.get("/admin/service-access")
@require_role("admin")
def service_access() -> ResponseReturnValue:
    conn = get_db()
    services = conn.execute("SELECT id, name FROM services ORDER BY name").fetchall()
    users = conn.execute("SELECT id, username, role, is_active FROM users ORDER BY username").fetchall()
    members = conn.execute(
        "SELECT user_id, service_id, role FROM service_members ORDER BY service_id, user_id"
    ).fetchall()
    memberships = {(row["user_id"], row["service_id"]): row["role"] for row in members}
    return Response(
        render_template("service_access.html", services=services, users=users, memberships=memberships, roles=("viewer", "editor", "service_admin")),
        headers={"Cache-Control": "no-store, private"},
    )


@routes.post("/admin/service-access/<int:service_id>/<int:user_id>")
@require_role("admin")
def service_access_grant(service_id: int, user_id: int) -> ResponseReturnValue:
    require_recent_reauth()
    role = request.form.get("role", "")
    if role not in SERVICE_ROLE_RANK:
        return Response("Papel inválido", status=400)
    conn = get_db()
    try:
        with transaction(conn):
            grant_membership(conn, _principal(), service_id=service_id, user_id=user_id, role=role)
    except NotFoundError:
        abort(404)
    except DomainError as error:
        return Response(error.message, status=400)
    return Response(status=204)


@routes.post("/admin/service-access/<int:service_id>/<int:user_id>/delete")
@require_role("admin")
def service_access_revoke(service_id: int, user_id: int) -> ResponseReturnValue:
    require_recent_reauth()
    conn = get_db()
    try:
        with transaction(conn):
            revoke_membership(conn, _principal(), service_id=service_id, user_id=user_id)
    except NotFoundError:
        abort(404)
    return Response(status=204)


def _webhook_form() -> tuple[str, str, bool, list[str]]:
    url = (request.form.get("url") or "").strip()
    description = (request.form.get("description") or "").strip()
    enabled = request.form.get("enabled") in {"1", "true", "on"}
    event_types = request.form.getlist("event_types")
    return url, description, enabled, event_types


@routes.get("/admin/security-integrations")
@require_role("admin")
def security_integrations() -> ResponseReturnValue:
    conn = get_db()
    configs = list_webhook_configs(conn)
    return Response(
        render_template(
            "security_integrations.html",
            configs=configs,
            event_types=webhook_event_types(),
            at_capacity=len(configs) >= 20,
        ),
        headers={"Cache-Control": "no-store, private"},
    )


@routes.post("/admin/security-integrations")
@require_role("admin")
def security_integration_create() -> ResponseReturnValue:
    require_recent_reauth()
    url, description, enabled, event_types = _webhook_form()
    conn = get_db()
    try:
        with transaction(conn):
            config_id, host, secret, subscriptions = create_webhook(
                conn,
                _principal(),
                url=url,
                description=description,
                enabled=enabled,
                event_types=event_types,
                data_key_b64=current_app.config["DATA_KEY_V1"],
                resolver=_webhook_resolver(),
            )
    except DomainError:
        return Response("Integração inválida", status=400)
    return Response(
        jsonify({"id": config_id, "signing_secret": secret}).get_data(as_text=True),
        status=201,
        headers={"Cache-Control": "no-store, private", "Content-Type": "application/json"},
    )


@routes.post("/admin/security-integrations/<int:config_id>")
@require_role("admin")
def security_integration_update(config_id: int) -> ResponseReturnValue:
    require_recent_reauth()
    url, description, enabled, event_types = _webhook_form()
    conn = get_db()
    try:
        with transaction(conn):
            update_webhook(
                conn,
                _principal(),
                config_id=config_id,
                url=url,
                description=description,
                enabled=enabled,
                event_types=event_types,
                data_key_b64=current_app.config["DATA_KEY_V1"],
                resolver=_webhook_resolver(),
            )
    except NotFoundError:
        abort(404)
    except DomainError:
        return Response("Integração inválida", status=400)
    return Response(status=204)


@routes.post("/admin/security-integrations/<int:config_id>/delete")
@require_role("admin")
def security_integration_delete(config_id: int) -> ResponseReturnValue:
    require_recent_reauth()
    conn = get_db()
    try:
        with transaction(conn):
            delete_webhook(conn, _principal(), config_id=config_id)
    except NotFoundError:
        abort(404)
    return Response(status=204)


@routes.post("/admin/security-integrations/<int:config_id>/test")
@require_role("admin")
def security_integration_test(config_id: int) -> ResponseReturnValue:
    require_recent_reauth()
    conn = get_db()
    try:
        with transaction(conn):
            test_webhook(conn, _principal(), config_id=config_id)
    except NotFoundError:
        abort(404)
    return Response(status=204)


@routes.get("/admin/settings")
@require_role("admin")
def settings_view() -> ResponseReturnValue:
    conn = get_db()
    feedback = OK_MESSAGES.get(request.args.get("ok") or "")
    return Response(
        render_template("settings.html", rotation_enabled=rotation_enabled(conn), feedback=feedback),
        headers={"Cache-Control": "no-store, private"},
    )


@routes.post("/admin/settings")
@require_role("admin")
def settings_update() -> ResponseReturnValue:
    require_recent_reauth()
    conn = get_db()
    enabled = request.form.get("rotation_enabled") in {"1", "true", "on"}
    with transaction(conn):
        set_rotation_enabled_setting(conn, _principal(), enabled=enabled)
    return redirect(url_for("routes.settings_view", ok="settings_updated"))


@routes.get("/admin/api-keys")
@require_role("admin")
def api_keys_view() -> ResponseReturnValue:
    conn = get_db()
    rows = list_api_keys(conn)
    feedback = OK_MESSAGES.get(request.args.get("ok") or "")
    return Response(
        render_template("api_keys.html", api_keys=rows, feedback=feedback),
        headers={"Cache-Control": "no-store, private"},
    )


@routes.post("/admin/api-keys")
@require_role("admin")
def api_keys_create() -> ResponseReturnValue:
    require_recent_reauth()
    name = (request.get_json(silent=True) or {}).get("name") if request.is_json else request.form.get("name")
    if name is None:
        name = request.form.get("name", "")
    conn = get_db()
    try:
        with transaction(conn):
            key_id, token = create_api_key(conn, principal_from_user(g.current_user), str(name or ""))
    except DomainError as error:
        return Response(error.message, status=error.status)
    return Response(
        jsonify({"id": key_id, "api_key": token}).get_data(as_text=True),
        status=201,
        headers={"Cache-Control": "no-store, private", "Content-Type": "application/json"},
    )


@routes.post("/admin/api-keys/<int:key_id>/revoke")
@require_role("admin")
def api_keys_revoke(key_id: int) -> ResponseReturnValue:
    require_recent_reauth()
    conn = get_db()
    try:
        with transaction(conn):
            revoke_api_key(conn, principal_from_user(g.current_user), key_id)
    except NotFoundError:
        abort(404)
    except DomainError as error:
        return Response(error.message, status=error.status)
    return Response(status=204)


@routes.post("/admin/api-keys/<int:key_id>/delete")
@require_role("admin")
def api_keys_delete(key_id: int) -> ResponseReturnValue:
    require_recent_reauth()
    conn = get_db()
    try:
        with transaction(conn):
            delete_api_key(conn, principal_from_user(g.current_user), key_id)
    except NotFoundError:
        abort(404)
    return redirect(url_for("routes.api_keys_view", ok="api_key_deleted"))

