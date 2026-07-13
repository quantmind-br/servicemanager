from __future__ import annotations

import csv
import io
import itertools
import zipfile
import xml.etree.ElementTree as ElementTree
from collections.abc import Iterator
from pathlib import PurePosixPath, PureWindowsPath
from typing import BinaryIO, Final

MAX_RECORDS: Final = 5_000
MAX_COLUMNS: Final = 20
MAX_CELLS: Final = 100_000
MAX_CELL_LENGTH: Final = 4_096
MAX_XLSX_MEMBERS: Final = 100
MAX_XLSX_UNCOMPRESSED_BYTES: Final = 25 * 1024 * 1024
MAX_XLSX_EXPANSION_RATIO: Final = 20


_ALLOWED_UPLOAD_MIME_TYPES: Final = {
    ".csv": frozenset({"text/csv", "application/csv", "application/vnd.ms-excel", "application/octet-stream"}),
    ".xlsx": frozenset({"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/octet-stream"}),
}


class ImportFormatError(ValueError):
    """A safe, classified import failure which never contains uploaded content."""

    def __init__(self, kind: str, message: str):
        super().__init__(message)
        self.kind = kind


def has_allowed_upload_mimetype(filename: str, mimetype: str | None) -> bool:
    """Treat Content-Type as an auxiliary signal; bytes still receive strict parsing."""
    normalized_name = filename.strip().lower()
    normalized_mimetype = (mimetype or "").split(";", 1)[0].strip().lower()
    return any(normalized_name.endswith(extension) and normalized_mimetype in allowed for extension, allowed in _ALLOWED_UPLOAD_MIME_TYPES.items())


def _as_binary_stream(data: bytes | BinaryIO) -> BinaryIO:
    if isinstance(data, bytes):
        return io.BytesIO(data)
    return data


def _checked_cell(value: object) -> str:
    cell = "" if value is None else str(value)
    if len(cell) > MAX_CELL_LENGTH:
        raise ImportFormatError("limits", "cell exceeds maximum length")
    return cell


def _validate_row_shape(row: list[str], *, cells: int) -> int:
    if len(row) > MAX_COLUMNS:
        raise ImportFormatError("limits", "row exceeds maximum columns")
    cells += len(row)
    if cells > MAX_CELLS:
        raise ImportFormatError("limits", "file exceeds maximum cells")
    return cells


def _csv_rows(stream: BinaryIO) -> Iterator[list[str]]:
    try:
        text = io.TextIOWrapper(stream, encoding="utf-8-sig", newline="")
        first_line = text.readline()
        delimiter = ";" if first_line.count(";") > first_line.count(",") else ","
        reader = csv.reader(itertools.chain((first_line,), text), delimiter=delimiter, strict=True)
        cells = 0
        for raw_row in reader:
            row = [_checked_cell(value) for value in raw_row]
            cells = _validate_row_shape(row, cells=cells)
            yield row
    except UnicodeDecodeError as error:
        raise ImportFormatError("format", "invalid UTF-8 CSV") from error
    except csv.Error as error:
        raise ImportFormatError("format", "invalid CSV") from error


def _validate_member_limits(member: zipfile.ZipInfo, *, total_size: int) -> int:
    total_size += member.file_size
    if total_size > MAX_XLSX_UNCOMPRESSED_BYTES:
        raise ImportFormatError("limits", "XLSX is too large when unpacked")
    if member.file_size and (
        member.compress_size == 0
        or member.file_size > member.compress_size * MAX_XLSX_EXPANSION_RATIO
    ):
        raise ImportFormatError("limits", "XLSX expansion ratio exceeds limit")
    return total_size


def _xml_attributes(archive: zipfile.ZipFile, member: zipfile.ZipInfo, *, kind: str) -> Iterator[dict[str, str]]:
    try:
        with archive.open(member) as xml_file:
            for _, element in ElementTree.iterparse(xml_file, events=("start",)):
                yield element.attrib
    except (ElementTree.ParseError, NotImplementedError, OSError, RuntimeError, zipfile.BadZipFile) as error:
        raise ImportFormatError("format", f"invalid XLSX {kind} XML") from error


def _reject_macro_content_types(archive: zipfile.ZipFile, member: zipfile.ZipInfo) -> None:
    for attributes in _xml_attributes(archive, member, kind="content types"):
        content_type = attributes.get("ContentType", "").casefold()
        if "vba" in content_type or "macroenabled" in content_type or "macrosheet" in content_type:
            raise ImportFormatError("format", "XLSX macros or external links are not allowed")


def _validate_relationships(archive: zipfile.ZipFile, member: zipfile.ZipInfo) -> None:
    for attributes in _xml_attributes(archive, member, kind="relationship"):
        relationship_type = attributes.get("Type", "").casefold()
        if "vba" in relationship_type or "macro" in relationship_type:
            raise ImportFormatError("format", "XLSX macros or external links are not allowed")
        if attributes.get("TargetMode") == "External":
            raise ImportFormatError("format", "XLSX external relationships are not allowed")

def _validate_xlsx_archive(stream: BinaryIO) -> None:
    try:
        stream.seek(0)
        with zipfile.ZipFile(stream) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_MEMBERS:
                raise ImportFormatError("limits", "XLSX has too many members")
            total_size = 0
            relationship_members: list[zipfile.ZipInfo] = []
            content_type_members: list[zipfile.ZipInfo] = []
            for member in members:
                path = PurePosixPath(member.filename)
                windows_path = PureWindowsPath(member.filename)
                if path.is_absolute() or windows_path.is_absolute() or ".." in path.parts or "\\" in member.filename:
                    raise ImportFormatError("format", "unsafe XLSX member path")
                normalized_name = member.filename.lower()
                if "vbaproject" in normalized_name or "/externallinks/" in f"/{normalized_name}":
                    raise ImportFormatError("format", "XLSX macros or external links are not allowed")
                total_size = _validate_member_limits(member, total_size=total_size)
                if normalized_name.endswith(".rels"):
                    relationship_members.append(member)
                elif normalized_name == "[content_types].xml":
                    content_type_members.append(member)
            for member in content_type_members:
                _reject_macro_content_types(archive, member)
            for member in relationship_members:
                _validate_relationships(archive, member)
    except ImportFormatError:
        raise
    except (OSError, zipfile.BadZipFile) as error:
        raise ImportFormatError("format", "invalid XLSX archive") from error


def _xlsx_rows(stream: BinaryIO) -> Iterator[list[str]]:
    _validate_xlsx_archive(stream)
    try:
        from openpyxl import load_workbook

        stream.seek(0)
        workbook = load_workbook(stream, read_only=True, data_only=True, keep_links=False)
        try:
            worksheet = workbook.active
            cells = 0
            for raw_row in worksheet.iter_rows(values_only=True):
                row = [_checked_cell(value) for value in raw_row]
                cells = _validate_row_shape(row, cells=cells)
                yield row
        finally:
            workbook.close()
    except ImportFormatError:
        raise
    except Exception as error:
        raise ImportFormatError("format", "invalid XLSX workbook") from error


def _import_rows(filename: str, stream: BinaryIO) -> Iterator[list[str]]:
    normalized_name = filename.strip().lower()
    if normalized_name.endswith(".csv"):
        yield from _csv_rows(stream)
        return
    if normalized_name.endswith(".xlsx"):
        yield from _xlsx_rows(stream)
        return
    raise ImportFormatError("format", "unsupported import format")


def parse_import_file(filename: str, data: bytes | BinaryIO) -> list[tuple[str, str, str]]:
    """Strictly parse a supported import stream without accepting lossy encodings."""
    rows = (row for row in _import_rows(filename, _as_binary_stream(data)) if any(row))
    try:
        header = next(rows)
    except StopIteration:
        return []
    headers = [value.strip().lower() for value in header]
    labelled = any(value in {"email", "e-mail", "senha", "password", "status"} for value in headers)
    positions = {
        "email": next((index for index, value in enumerate(headers) if value in {"email", "e-mail"}), None if labelled else 0),
        "password": next((index for index, value in enumerate(headers) if value in {"password", "senha", "pass"}), None if labelled else 1),
        "status": next((index for index, value in enumerate(headers) if value == "status"), None if labelled else 2),
    }
    if labelled and positions["email"] is None:
        raise ImportFormatError("format", "labeled import requires an email column")

    records: list[tuple[str, str, str]] = []
    body = rows if labelled else itertools.chain((header,), rows)
    for row in body:
        if len(records) >= MAX_RECORDS:
            raise ImportFormatError("limits", "file exceeds maximum records")

        def cell(index: int | None) -> str:
            return row[index] if index is not None and index < len(row) else ""

        records.append((cell(positions["email"]), cell(positions["password"]), cell(positions["status"])))
    return records
